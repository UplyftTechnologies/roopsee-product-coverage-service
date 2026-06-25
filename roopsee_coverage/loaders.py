from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .constants import EYES_SHEET, FACE_SHEET, LIPS_SHEET
from .models import ScoreRow
from .utils import clean_text, first_image, norm_key, safe_float


CATALOG_SCORE_COLUMNS = [
    "<16",
    "17-25",
    "Above 25",
    "Acne",
    "Body Acne",
    "Dryness",
    "Open Pores",
    "Uneven Skin Tone",
    "Dark Spots/Pigmentation",
    "Melasma",
    "Barrier Repair",
    "Comedones",
    "Wrinkles/Fine lines",
    "Redness/Irritation",
    "Dehydration",
    "Dullness",
    "Tanning",
    "Concern weight",
    "Oily Score",
    "Oily+Sensitive Score",
    "Dry Score",
    "Dry+Sensitive Score",
    "Normal Score",
    "Normal+Sensitive Score",
    "Combination Score",
    "Combination+Sensitive Score",
    "Excessive Dryness score",
    "Pregnancy Score",
    "Breastfeeling Score",
]


def parse_catalog(products_csv: Path) -> dict[str, dict[str, Any]]:
    products: dict[str, dict[str, Any]] = {}
    with products_csv.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            uid = clean_text(row.get("product_uid"))
            if not uid:
                continue
            products[norm_key(uid)] = {
                "product_uid": uid,
                "product_name": clean_text(row.get("product_name")),
                "brand_name": clean_text(row.get("brand_name")),
                "category": clean_text(row.get("category")),
                "product_type": clean_text(row.get("product_type")),
                "addresses_skin_concerns": clean_text(row.get("addresses_skin_concerns")),
                "sku_size": clean_text(row.get("sku_size")),
                "mrp": clean_text(row.get("mrp")),
                "sp": clean_text(row.get("sp")),
                "single_hero_ingredient": clean_text(row.get("single_hero_ingredient")),
                "secondary_hero_ingredients": clean_text(row.get("secondary_hero_ingredients")),
                "dos": clean_text(row.get("dos")),
                "donts": clean_text(row.get("donts")),
                "storage_instructions": clean_text(row.get("storage_instructions")),
                "usage_instructions": clean_text(row.get("usage_instructions")),
                "when_to_use": clean_text(row.get("when_to_use")),
                "ingredient_cautions": clean_text(row.get("ingredient_cautions")),
                "product_description": clean_text(row.get("product_description")),
                "ingredients": clean_text(row.get("ingredients")),
                "image": first_image(row.get("images", "")),
                "database_id": clean_text(row.get("id")),
            }
    return products


def parse_catalog_score_rows(products_csv: Path) -> list[ScoreRow]:
    rows: list[ScoreRow] = []
    with products_csv.open(newline="", encoding="utf-8-sig") as handle:
        for source_row, row in enumerate(csv.DictReader(handle), start=2):
            uid = clean_text(row.get("product_uid"))
            name = clean_text(row.get("product_name"))
            if not uid or not name:
                continue

            scores: dict[str, float] = {}
            for header in CATALOG_SCORE_COLUMNS:
                score = safe_float(row.get(header))
                if score is not None:
                    scores[header] = score

            above_25_score = scores.get("Above 25")
            if above_25_score is not None and "+>25" not in scores:
                scores["+>25"] = above_25_score

            if not scores:
                continue

            rows.append(
                ScoreRow(
                    source_sheet=FACE_SHEET,
                    product_uid=uid,
                    product_name=name,
                    brand=clean_text(row.get("brand_name")),
                    hero_ingredient=clean_text(row.get("single_hero_ingredient")),
                    secondary_ingredients=clean_text(row.get("secondary_hero_ingredients")),
                    category=clean_text(row.get("category")),
                    product_type=clean_text(row.get("product_type")),
                    scores=scores,
                    source_row=source_row,
                )
            )
    return rows


def parse_score_sheet(ws: Any, source_sheet: str, header_row: int, data_start_row: int) -> list[ScoreRow]:
    headers = [clean_text(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[ScoreRow] = []
    for row_number in range(data_start_row, ws.max_row + 1):
        values = [ws.cell(row_number, col).value for col in range(1, ws.max_column + 1)]
        uid = clean_text(values[0] if len(values) > 0 else "")
        name = clean_text(values[1] if len(values) > 1 else "")
        if not uid or not name:
            continue

        scores: dict[str, float] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            score = safe_float(values[idx] if idx < len(values) else None)
            if score is not None:
                scores[header] = score

        rows.append(
            ScoreRow(
                source_sheet=source_sheet,
                product_uid=uid,
                product_name=name,
                brand=clean_text(values[2] if len(values) > 2 else ""),
                hero_ingredient=clean_text(values[3] if len(values) > 3 else ""),
                secondary_ingredients=clean_text(values[4] if len(values) > 4 else ""),
                category=clean_text(values[5] if len(values) > 5 else ""),
                product_type=clean_text(values[6] if len(values) > 6 else ""),
                scores=scores,
                source_row=row_number,
            )
        )
    return rows


def load_score_rows(score_workbook: Path, products_csv: Path | None = None) -> list[ScoreRow]:
    workbook = load_workbook(score_workbook, data_only=True, read_only=False)
    workbook_rows: list[ScoreRow] = []
    workbook_rows.extend(parse_score_sheet(workbook[FACE_SHEET], FACE_SHEET, header_row=2, data_start_row=3))
    workbook_rows.extend(parse_score_sheet(workbook[LIPS_SHEET], LIPS_SHEET, header_row=1, data_start_row=2))
    workbook_rows.extend(parse_score_sheet(workbook[EYES_SHEET], EYES_SHEET, header_row=1, data_start_row=2))
    catalog_rows: list[ScoreRow] = []
    if products_csv is not None:
        catalog_rows = parse_catalog_score_rows(products_csv)
    catalog_keys = {norm_key(row.product_uid) for row in catalog_rows}
    rows = [row for row in workbook_rows if norm_key(row.product_uid) not in catalog_keys]
    rows.extend(catalog_rows)
    return rows
