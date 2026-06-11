from __future__ import annotations

import csv
import json
import os
import re
from dataclasses import dataclass
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"

DEFAULT_SCORE_WORKBOOK = str(BASE_DIR / "data" / "Product details and score logic.xlsx")
DEFAULT_PRODUCTS_CSV = str(BASE_DIR / "data" / "products.csv")

FACE_SHEET = "Face and body"
LIPS_SHEET = "Lips"
EYES_SHEET = "Eyes"

QUIZ_OPTIONS = {
    "skinTypes": ["Oily", "Dry", "Normal", "Combination"],
    "faceBodyConcerns": [
        "Acne",
        "Pigmentation",
        "Dryness",
        "Dark Spots",
        "Aging",
        "Sensitivity",
        "Large Pores",
        "Dullness",
    ],
    "lipsEyesConcerns": [
        "Dark circles",
        "Puffiness",
        "Dry Under Eye",
        "Sensitive Eye",
        "Chapped Lips",
        "Lip Pigment",
        "Dull Lips",
        "Dehydrated Lips",
    ],
    "specialConditions": ["Excessive Dryness", "Pregnant", "Breastfeeding", "None"],
    "ages": ["Under 16", "17-25", "Above 25"],
    "genders": ["male", "female", "other", "prefer not to say"],
}

AGE_COLUMN_MAP = {
    "under 16": "<16",
    "below 16": "<16",
    "<16": "<16",
    "16": "<16",
    "17-25": "17-25",
    "17 - 25": "17-25",
    "17 25": "17-25",
    "above 25": "+>25",
    "over 25": "+>25",
    "25+": "+>25",
    "+>25": "+>25",
}

FACE_CONCERN_MAP = {
    "acne": ["Acne", "Comedones", "Body Acne"],
    "pigmentation": ["Dark Spots/Pigmentation", "Uneven Skin Tone", "Melasma"],
    "dryness": ["Dryness", "Dehydration", "Barrier Repair"],
    "dark spots": ["Dark Spots/Pigmentation", "Melasma", "Uneven Skin Tone"],
    "aging": ["Wrinkles/Fine lines"],
    "sensitivity": ["Redness/Irritation"],
    "large pores": ["Open Pores"],
    "dullness": ["Dullness", "Uneven Skin Tone"],
}

LIP_CONCERN_MAP = {
    "chapped lips": ["Dry Lips/Chapped Lips"],
    "dry lips": ["Dry Lips/Chapped Lips"],
    "lip pigment": ["Lip Pigmentation"],
    "lip pigmentation": ["Lip Pigmentation"],
    "dull lips": ["Dull Looking lips"],
    "dehydrated lips": ["Dehydrated Lips"],
}

EYE_CONCERN_MAP = {
    "dark circles": ["Dark circles"],
    "puffiness": ["Puffiness/Eye Bags"],
    "dry under eye": ["Dry/Dehydrated Under Eyes"],
    "dry under eyes": ["Dry/Dehydrated Under Eyes"],
    "sensitive eye": ["Sensitive/Irritated Eye area"],
    "sensitive eyes": ["Sensitive/Irritated Eye area"],
}

SCORE_LABELS = [
    (90, "Excellent Match"),
    (80, "Great Match"),
    (70, "Good Match"),
    (50, "Fits with Caution"),
    (-999, "Not Recommended"),
]


@dataclass
class ScoreRow:
    source_sheet: str
    product_uid: str
    product_name: str
    brand: str
    category: str
    product_type: str
    hero_ingredient: str
    secondary_ingredients: str
    scores: dict[str, float]
    source_row: int


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split()).strip()


def norm_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())


def norm_label(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean_text(value).lower()).strip()


def safe_float(value: Any) -> float | None:
    if value is None or clean_text(value) == "":
        return None
    try:
        return float(str(value).replace("%", "").strip())
    except ValueError:
        return None


def first_image(raw: str) -> str:
    raw = clean_text(raw)
    if not raw:
        return ""
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, list) and parsed:
            return clean_text(parsed[0])
    except json.JSONDecodeError:
        pass
    return raw


def parse_catalog(products_csv: Path) -> dict[str, dict[str, Any]]:
    products: dict[str, dict[str, Any]] = {}
    with products_csv.open(newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            uid = clean_text(row.get("product_uid"))
            if not uid:
                continue
            key = norm_key(uid)
            products[key] = {
                "product_uid": uid,
                "product_name": clean_text(row.get("product_name")),
                "brand_name": clean_text(row.get("brand_name")),
                "category": clean_text(row.get("category")),
                "product_type": clean_text(row.get("product_type")),
                "addresses_skin_concerns": clean_text(row.get("addresses_skin_concerns")),
                "sku_size": clean_text(row.get("sku_size")),
                "mrp": clean_text(row.get("mrp")),
                "sp": clean_text(row.get("sp")),
                "single_hero_ingredient": clean_text(row.get("single_hero_ingredient")),
                "secondary_hero_ingredients": clean_text(row.get("secondary_hero_ingredients")),
                "when_to_use": clean_text(row.get("when_to_use")),
                "ingredient_cautions": clean_text(row.get("ingredient_cautions")),
                "product_description": clean_text(row.get("product_description")),
                "image": first_image(row.get("images", "")),
                "database_id": clean_text(row.get("id")),
            }
    return products


def parse_score_sheet(ws: Any, source_sheet: str, header_row: int, data_start_row: int) -> list[ScoreRow]:
    headers = [clean_text(ws.cell(header_row, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[ScoreRow] = []
    for row_number in range(data_start_row, ws.max_row + 1):
        values = [ws.cell(row_number, col).value for col in range(1, ws.max_column + 1)]
        uid = clean_text(values[0] if len(values) > 0 else "")
        name = clean_text(values[1] if len(values) > 1 else "")
        if not uid or not name:
            continue

        scores: dict[str, float] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            score = safe_float(values[idx] if idx < len(values) else None)
            if score is not None:
                scores[header] = score

        rows.append(
            ScoreRow(
                source_sheet=source_sheet,
                product_uid=uid,
                product_name=name,
                brand=clean_text(values[2] if len(values) > 2 else ""),
                hero_ingredient=clean_text(values[3] if len(values) > 3 else ""),
                secondary_ingredients=clean_text(values[4] if len(values) > 4 else ""),
                category=clean_text(values[5] if len(values) > 5 else ""),
                product_type=clean_text(values[6] if len(values) > 6 else ""),
                scores=scores,
                source_row=row_number,
            )
        )
    return rows


def load_score_rows(score_workbook: Path) -> list[ScoreRow]:
    workbook = load_workbook(score_workbook, data_only=True, read_only=False)
    rows: list[ScoreRow] = []
    rows.extend(parse_score_sheet(workbook[FACE_SHEET], FACE_SHEET, header_row=2, data_start_row=3))
    rows.extend(parse_score_sheet(workbook[LIPS_SHEET], LIPS_SHEET, header_row=1, data_start_row=2))
    rows.extend(parse_score_sheet(workbook[EYES_SHEET], EYES_SHEET, header_row=1, data_start_row=2))
    return rows


def age_column(age: str) -> str:
    return AGE_COLUMN_MAP.get(norm_label(age), "+>25")


def is_sensitive_profile(profile: dict[str, Any]) -> bool:
    face_concerns = [norm_label(item) for item in profile.get("selectedFaceBodyConcerns", [])]
    lips_eye = [norm_label(item) for item in profile.get("selectedLipsEyesConcerns", [])]
    return any(item in {"sensitivity", "sensitive eye", "sensitive eyes"} for item in face_concerns + lips_eye)


def skin_column(skin_type: str, sensitive: bool) -> str:
    base = clean_text(skin_type).title()
    if base not in {"Oily", "Dry", "Normal", "Combination"}:
        base = "Normal"
    return f"{base}+Sensitive Score" if sensitive else f"{base} Score"


def available_scores(row: ScoreRow, headers: list[str]) -> list[float]:
    values = [row.scores[header] for header in headers if header in row.scores]
    return values


def map_special_columns(row: ScoreRow, special_conditions: list[str]) -> list[tuple[str, float]]:
    conditions = [norm_label(item) for item in special_conditions if norm_label(item)]
    if not conditions or conditions == ["none"] or "none" in conditions and len(conditions) == 1:
        none_value = row.scores.get("None")
        return [("None", none_value)] if none_value is not None else []

    mapped: list[tuple[str, float]] = []
    if "pregnant" in conditions:
        value = row.scores.get("Pregnancy Score", row.scores.get("Pregnancy safe"))
        if value is not None:
            mapped.append(("Pregnancy", value))
    if "breastfeeding" in conditions:
        value = row.scores.get("Breastfeeling Score", row.scores.get("Breastfeeding safe"))
        if value is not None:
            mapped.append(("Breastfeeding", value))
    if "excessive dryness" in conditions:
        value = row.scores.get("Excessive Dryness score")
        if value is None and row.source_sheet == LIPS_SHEET:
            dry_values = available_scores(row, ["Dry Lips/Chapped Lips", "Dehydrated Lips"])
            value = max(dry_values) if dry_values else None
        if value is None and row.source_sheet == EYES_SHEET:
            dry_values = available_scores(row, ["Dry/Dehydrated Under Eyes"])
            value = max(dry_values) if dry_values else None
        if value is not None:
            mapped.append(("Excessive Dryness", value))
    return mapped


def concern_headers_for_row(row: ScoreRow, profile: dict[str, Any]) -> list[tuple[str, list[str]]]:
    mapped: list[tuple[str, list[str]]] = []
    if row.source_sheet == FACE_SHEET:
        for concern in profile.get("selectedFaceBodyConcerns", []):
            headers = FACE_CONCERN_MAP.get(norm_label(concern), [clean_text(concern)])
            mapped.append((clean_text(concern), headers))
    elif row.source_sheet == LIPS_SHEET:
        for concern in profile.get("selectedLipsEyesConcerns", []):
            headers = LIP_CONCERN_MAP.get(norm_label(concern))
            if headers:
                mapped.append((clean_text(concern), headers))
    elif row.source_sheet == EYES_SHEET:
        for concern in profile.get("selectedLipsEyesConcerns", []):
            headers = EYE_CONCERN_MAP.get(norm_label(concern))
            if headers:
                mapped.append((clean_text(concern), headers))
    return mapped


def target_sheets(profile: dict[str, Any]) -> set[str]:
    sheets: set[str] = set()
    if profile.get("selectedFaceBodyConcerns"):
        sheets.add(FACE_SHEET)
    for concern in profile.get("selectedLipsEyesConcerns", []):
        key = norm_label(concern)
        if key in LIP_CONCERN_MAP:
            sheets.add(LIPS_SHEET)
        if key in EYE_CONCERN_MAP:
            sheets.add(EYES_SHEET)
    if not sheets:
        sheets.add(FACE_SHEET)
    return sheets


def label_for_score(score: float) -> str:
    for threshold, label in SCORE_LABELS:
        if score >= threshold:
            return label
    return "Not Recommended"


def weighted_average(components: list[tuple[str, float, float]]) -> float:
    total_weight = sum(weight for _, _, weight in components if weight > 0)
    if total_weight <= 0:
        return 0.0
    return sum(value * weight for _, value, weight in components if weight > 0) / total_weight


def score_row_for_profile(row: ScoreRow, catalog: dict[str, Any], profile: dict[str, Any]) -> dict[str, Any]:
    components: list[tuple[str, float, float]] = []
    reasons: list[str] = []
    warnings: list[str] = []

    age_header = age_column(clean_text(profile.get("age", "")))
    age_score = row.scores.get(age_header)
    if age_score is not None:
        components.append((f"Age {age_header}", age_score, 0.10 if row.source_sheet == FACE_SHEET else 0.15))
        reasons.append(f"Age fit {age_header}: {age_score:g}")

    concern_pairs = concern_headers_for_row(row, profile)
    concern_scores: list[float] = []
    for concern, headers in concern_pairs:
        values = available_scores(row, headers)
        if values:
            best_value = max(values)
            concern_scores.append(best_value)
            reasons.append(f"{concern}: {best_value:g}")
    if concern_scores:
        concern_weight = 0.45 if row.source_sheet == FACE_SHEET else 0.60
        components.append(("Concern fit", sum(concern_scores) / len(concern_scores), concern_weight))
    elif row.scores.get("None") is not None:
        components.append(("General fit", row.scores["None"], 0.25))

    if row.source_sheet == FACE_SHEET:
        skin_header = skin_column(clean_text(profile.get("selectedSkinType", "Normal")), is_sensitive_profile(profile))
        skin_score = row.scores.get(skin_header)
        if skin_score is not None:
            components.append((skin_header, skin_score, 0.30))
            reasons.append(f"{skin_header}: {skin_score:g}")

    special_pairs = map_special_columns(row, profile.get("selectedSpecialConditions", []))
    special_values = [value for _, value in special_pairs]
    if special_values:
        special_score = min(special_values) if any(value < 0 for value in special_values) else sum(special_values) / len(special_values)
        components.append(("Special conditions", special_score, 0.15 if row.source_sheet == FACE_SHEET else 0.25))
        for label, value in special_pairs:
            reasons.append(f"{label}: {value:g}")
        if any(value < 0 for value in special_values):
            warnings.append("Safety conflict with selected pregnancy/breastfeeding condition.")

    raw_score = weighted_average(components)
    if any(value < 0 for value in special_values):
        raw_score = min(raw_score, 30)

    final_score = max(0, min(100, round(raw_score, 1)))
    label = label_for_score(final_score)
    explanation = (
        f"{label}: ranked from doctor-sheet scores for "
        f"{', '.join(reason for reason in reasons[:5])}. "
        "Only products present in the live catalog CSV are returned."
    )
    if warnings:
        explanation = f"{warnings[0]} {explanation}"

    return {
        "product_uid": catalog["product_uid"],
        "score_uid": row.product_uid,
        "product_name": catalog["product_name"] or row.product_name,
        "brand_name": catalog["brand_name"] or row.brand,
        "category": catalog["category"] or row.category,
        "product_type": catalog["product_type"] or row.product_type,
        "score": final_score,
        "match_label": label,
        "explanation": explanation,
        "source_sheet": row.source_sheet,
        "hero_ingredient": catalog["single_hero_ingredient"] or row.hero_ingredient,
        "secondary_hero_ingredients": catalog["secondary_hero_ingredients"] or row.secondary_ingredients,
        "size": catalog["sku_size"],
        "mrp": catalog["mrp"],
        "selling_price": catalog["sp"],
        "when_to_use": catalog["when_to_use"],
        "image": catalog["image"],
        "component_scores": [{"name": name, "score": round(value, 1), "weight": weight} for name, value, weight in components],
        "warnings": warnings,
    }


class RecommendationEngine:
    def __init__(self, score_workbook: Path, products_csv: Path):
        self.score_workbook = score_workbook
        self.products_csv = products_csv
        self.catalog = parse_catalog(products_csv)
        self.score_rows = load_score_rows(score_workbook)

        self.score_rows_by_catalog_key: dict[str, list[ScoreRow]] = {}
        self.score_only_uids: list[str] = []
        for row in self.score_rows:
            key = norm_key(row.product_uid)
            if key in self.catalog:
                self.score_rows_by_catalog_key.setdefault(key, []).append(row)
            else:
                self.score_only_uids.append(row.product_uid)

        self.catalog_missing_score = [
            product["product_uid"]
            for key, product in self.catalog.items()
            if key not in self.score_rows_by_catalog_key
        ]

    def recommend(self, profile: dict[str, Any], limit: int = 60) -> dict[str, Any]:
        wanted_sheets = target_sheets(profile)
        best_by_uid: dict[str, dict[str, Any]] = {}
        for key, rows in self.score_rows_by_catalog_key.items():
            catalog = self.catalog[key]
            for row in rows:
                if row.source_sheet not in wanted_sheets:
                    continue
                scored = score_row_for_profile(row, catalog, profile)
                existing = best_by_uid.get(catalog["product_uid"])
                if existing is None or scored["score"] > existing["score"]:
                    best_by_uid[catalog["product_uid"]] = scored

        results = sorted(
            best_by_uid.values(),
            key=lambda item: (-item["score"], item["category"], item["product_name"].lower()),
        )
        results = results[: max(1, min(limit, 250))]
        return {
            "profile": profile,
            "target_sheets": sorted(wanted_sheets),
            "total_matches": len(best_by_uid),
            "returned": len(results),
            "summary": summarize_results(results),
            "products": results,
        }

    def health(self) -> dict[str, Any]:
        by_sheet: dict[str, int] = {}
        for row in self.score_rows:
            by_sheet[row.source_sheet] = by_sheet.get(row.source_sheet, 0) + 1
        return {
            "score_workbook": str(self.score_workbook),
            "products_csv": str(self.products_csv),
            "catalog_products": len(self.catalog),
            "score_rows": len(self.score_rows),
            "score_rows_by_sheet": by_sheet,
            "catalog_missing_score_count": len(self.catalog_missing_score),
            "catalog_missing_score": self.catalog_missing_score,
            "score_only_count": len(self.score_only_uids),
            "score_only_uids": self.score_only_uids,
        }

    def coverage(self, profiles: list[dict[str, Any]], top_n: int = 12) -> dict[str, Any]:
        rows = []
        for index, profile in enumerate(profiles, start=1):
            response = self.recommend(profile, limit=250)
            summary = response["summary"]
            status = coverage_status(profile, summary)
            rows.append({
                "profile_id": f"profile_{index:03d}",
                "profile": profile,
                "status": status,
                "total_matches": response["total_matches"],
                "excellent_count": summary["excellent_count"],
                "great_or_better_count": summary["great_or_better_count"],
                "good_or_better_count": summary["good_or_better_count"],
                "top_products": response["products"][:top_n],
                "target_sheets": response["target_sheets"],
            })

        status_counts: dict[str, int] = {}
        for row in rows:
            status_counts[row["status"]] = status_counts.get(row["status"], 0) + 1

        return {
            "profile_count": len(rows),
            "status_counts": status_counts,
            "catalog_products": len(self.catalog),
            "catalog_missing_score_count": len(self.catalog_missing_score),
            "rows": rows,
        }


def summarize_results(results: list[dict[str, Any]]) -> dict[str, Any]:
    categories: dict[str, int] = {}
    product_types: dict[str, int] = {}
    for item in results:
        categories[item["category"]] = categories.get(item["category"], 0) + 1
        product_types[item["product_type"]] = product_types.get(item["product_type"], 0) + 1
    return {
        "excellent_count": sum(1 for item in results if item["score"] >= 90),
        "great_or_better_count": sum(1 for item in results if item["score"] >= 80),
        "good_or_better_count": sum(1 for item in results if item["score"] >= 70),
        "caution_or_better_count": sum(1 for item in results if item["score"] >= 50),
        "categories": categories,
        "product_types": product_types,
        "top_score": results[0]["score"] if results else None,
        "bottom_score": results[-1]["score"] if results else None,
    }


def coverage_status(profile: dict[str, Any], summary: dict[str, Any]) -> str:
    lips_eye_only = bool(profile.get("selectedLipsEyesConcerns")) and not bool(profile.get("selectedFaceBodyConcerns"))
    if lips_eye_only:
        if summary["good_or_better_count"] >= 4 and summary["great_or_better_count"] >= 2:
            return "Strong"
        if summary["good_or_better_count"] >= 2:
            return "Limited but usable"
        return "Coverage gap"

    if summary["good_or_better_count"] >= 12 and summary["great_or_better_count"] >= 5:
        return "Strong"
    if summary["good_or_better_count"] >= 8 and summary["great_or_better_count"] >= 3:
        return "Usable"
    return "Coverage gap"


def representative_profiles(limit: int = 72) -> list[dict[str, Any]]:
    skin_types = QUIZ_OPTIONS["skinTypes"]
    ages = QUIZ_OPTIONS["ages"]
    face_pairs = [
        ["Acne"],
        ["Dark Spots"],
        ["Dryness"],
        ["Pigmentation"],
        ["Aging"],
        ["Sensitivity"],
        ["Large Pores"],
        ["Dullness"],
        ["Acne", "Dark Spots"],
        ["Acne", "Large Pores"],
        ["Dryness", "Sensitivity"],
        ["Pigmentation", "Dullness"],
        ["Aging", "Dryness"],
        ["Dark Spots", "Dullness"],
    ]
    lips_eye_pairs = [
        ["Dark circles"],
        ["Puffiness"],
        ["Dry Under Eye"],
        ["Sensitive Eye"],
        ["Chapped Lips"],
        ["Lip Pigment"],
        ["Dull Lips"],
        ["Dehydrated Lips"],
        ["Dark circles", "Puffiness"],
        ["Chapped Lips", "Lip Pigment"],
        ["Dry Under Eye", "Sensitive Eye"],
        ["Dull Lips", "Dehydrated Lips"],
    ]
    specials = [
        ["None"],
        ["Excessive Dryness"],
        ["Pregnant"],
        ["Breastfeeding"],
        ["Pregnant", "Breastfeeding"],
    ]

    profiles: list[dict[str, Any]] = []
    for skin_index, skin_type in enumerate(skin_types):
        for concern_index, concerns in enumerate(face_pairs):
            profiles.append({
                "age": ages[(skin_index + concern_index) % len(ages)],
                "selectedGender": "male" if concern_index % 2 == 0 else "female",
                "selectedSkinType": skin_type,
                "selectedFaceBodyConcerns": concerns,
                "selectedLipsEyesConcerns": [],
                "selectedSpecialConditions": specials[(skin_index + concern_index) % len(specials)],
            })
            if len(profiles) >= limit:
                return profiles

    for skin_index, skin_type in enumerate(skin_types):
        for concern_index, concerns in enumerate(lips_eye_pairs):
            profiles.append({
                "age": ages[(skin_index + concern_index) % len(ages)],
                "selectedGender": "male" if skin_index % 2 == 0 else "female",
                "selectedSkinType": skin_type,
                "selectedFaceBodyConcerns": [],
                "selectedLipsEyesConcerns": concerns,
                "selectedSpecialConditions": specials[(skin_index + concern_index) % len(specials)],
            })
            if len(profiles) >= limit:
                return profiles

    return profiles[:limit]


def read_request_json(handler: BaseHTTPRequestHandler) -> dict[str, Any]:
    length = int(handler.headers.get("Content-Length", "0") or "0")
    if length <= 0:
        return {}
    raw = handler.rfile.read(length).decode("utf-8")
    return json.loads(raw or "{}")


def send_json(handler: BaseHTTPRequestHandler, payload: Any, status: int = 200) -> None:
    body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Access-Control-Allow-Origin", "*")
    handler.send_header("Access-Control-Allow-Headers", "Content-Type")
    handler.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def send_file(handler: BaseHTTPRequestHandler, path: Path) -> None:
    if not path.exists() or not path.is_file():
        handler.send_error(HTTPStatus.NOT_FOUND)
        return
    content_type = "text/html; charset=utf-8" if path.suffix == ".html" else "text/plain; charset=utf-8"
    body = path.read_bytes()
    handler.send_response(HTTPStatus.OK)
    handler.send_header("Content-Type", content_type)
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def make_handler(engine: RecommendationEngine):
    class Handler(BaseHTTPRequestHandler):
        def do_OPTIONS(self) -> None:
            send_json(self, {"ok": True})

        def do_GET(self) -> None:
            parsed = urlparse(self.path)
            query = parse_qs(parsed.query)
            try:
                if parsed.path in {"/", "/index.html"}:
                    send_file(self, STATIC_DIR / "index.html")
                elif parsed.path == "/api/health":
                    send_json(self, engine.health())
                elif parsed.path == "/api/options":
                    payload = {"quiz_options": QUIZ_OPTIONS, "health": engine.health()}
                    send_json(self, payload)
                elif parsed.path == "/api/representative-profiles":
                    count = int(query.get("count", ["72"])[0])
                    send_json(self, {"profiles": representative_profiles(count)})
                elif parsed.path == "/api/coverage":
                    count = int(query.get("count", ["72"])[0])
                    top_n = int(query.get("top_n", ["5"])[0])
                    send_json(self, engine.coverage(representative_profiles(count), top_n=top_n))
                else:
                    requested = (STATIC_DIR / parsed.path.lstrip("/")).resolve()
                    if STATIC_DIR in requested.parents:
                        send_file(self, requested)
                    else:
                        self.send_error(HTTPStatus.NOT_FOUND)
            except Exception as exc:  # noqa: BLE001
                send_json(self, {"error": str(exc)}, status=500)

        def do_POST(self) -> None:
            parsed = urlparse(self.path)
            query = parse_qs(parsed.query)
            try:
                body = read_request_json(self)
                if parsed.path == "/api/recommend":
                    limit = int(query.get("limit", [str(body.get("limit", 60))])[0])
                    send_json(self, engine.recommend(body, limit=limit))
                elif parsed.path == "/api/coverage":
                    profiles = body.get("profiles") or representative_profiles(int(body.get("count", 72)))
                    top_n = int(body.get("top_n", 5))
                    send_json(self, engine.coverage(profiles, top_n=top_n))
                else:
                    self.send_error(HTTPStatus.NOT_FOUND)
            except Exception as exc:  # noqa: BLE001
                send_json(self, {"error": str(exc)}, status=500)

        def log_message(self, format: str, *args: Any) -> None:
            print(f"{self.address_string()} - {format % args}")

    return Handler


def main() -> None:
    score_workbook = Path(os.getenv("SCORE_WORKBOOK_PATH", DEFAULT_SCORE_WORKBOOK)).expanduser()
    products_csv = Path(os.getenv("PRODUCTS_CSV_PATH", DEFAULT_PRODUCTS_CSV)).expanduser()
    port = int(os.getenv("PORT", "8020"))
    host = os.getenv("HOST", "127.0.0.1")

    engine = RecommendationEngine(score_workbook, products_csv)
    server = ThreadingHTTPServer((host, port), make_handler(engine))
    print(f"Roopsee product coverage service running at http://{host}:{port}")
    print(json.dumps(engine.health(), indent=2))
    server.serve_forever()


if __name__ == "__main__":
    main()
