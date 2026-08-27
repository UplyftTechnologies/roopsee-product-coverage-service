from __future__ import annotations

import csv
import difflib
import gzip
import json
import math
import os
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = REPO_ROOT / "data" / "source"


def env_path(name: str, default: Path) -> Path:
    return Path(os.environ.get(name, str(default))).expanduser()


USEFUL_PRODUCTS = env_path("ROOPSEE_USEFUL_PRODUCTS", SOURCE_DIR / "useful_skin_bodycare_products.xlsx")
RETAILER_PRODUCTS = env_path("ROOPSEE_RETAILER_PRODUCTS", SOURCE_DIR / "retailer_products_rows.csv.gz")
INGREDIENT_SCORES = env_path("ROOPSEE_INGREDIENT_SCORES", SOURCE_DIR / "roopsee_ingredient_scores_v3.xlsx")
OUTPUT_DIR = env_path("ROOPSEE_AUTO_OUTPUT_DIR", REPO_ROOT / "outputs" / "roopsee_automated_scoring")
PAYLOAD_PATH = OUTPUT_DIR / "automated_scoring_payload.json"


def open_text(path: Path):
    if path.suffix == ".gz":
        return gzip.open(path, mode="rt", newline="", encoding="utf-8-sig")
    return path.open(newline="", encoding="utf-8-sig")


AGE_COLUMNS = ["<16", "17-25", "+>25"]
CONCERN_COLUMNS = [
    "Acne",
    "Body Acne",
    "Dryness",
    "Open Pores",
    "Uneven Skin Tone",
    "Dark Spots/Pigmentation",
    "Melasma",
    "Barrier Repair",
    "Comedones",
    "Wrinkles/Fine lines",
    "Redness/Irritation",
    "Dehydration",
    "Dullness",
    "Tanning",
]
SKIN_COLUMNS = [
    "Oily Score",
    "Oily+Sensitive Score",
    "Dry Score",
    "Dry+Sensitive Score",
    "Normal Score",
    "Normal+Sensitive Score",
    "Combination Score",
    "Combination+Sensitive Score",
    "I dont know Score",
    "I dont know+Sensitive Score",
]
SPECIAL_COLUMNS = ["Excessive Dryness score", "Pregnancy Score", "Breastfeeling Score", "None"]
SCORE_COLUMNS = AGE_COLUMNS + CONCERN_COLUMNS + SKIN_COLUMNS + SPECIAL_COLUMNS

INGREDIENT_SCORE_HEADER_MAP = {
    "<16": "<16",
    "17-25": "17-25",
    "Above 25": "+>25",
    "Acne": "Acne",
    "Body Acne": "Body Acne",
    "Dryness": "Dryness",
    "Open Pores": "Open Pores",
    "Uneven Skin Tone": "Uneven Skin Tone",
    "Dark Spots/Pigmentation": "Dark Spots/Pigmentation",
    "Melasma": "Melasma",
    "Barrier Repair": "Barrier Repair",
    "Comedones": "Comedones",
    "Wrinkles/Fine lines": "Wrinkles/Fine lines",
    "Redness/Irritation": "Redness/Irritation",
    "Dehydration": "Dehydration",
    "Dullness": "Dullness",
    "Tanning": "Tanning",
    "Oily Score": "Oily Score",
    "Oily+Sensitive Score": "Oily+Sensitive Score",
    "Dry Score": "Dry Score",
    "Dry+Sensitive Score": "Dry+Sensitive Score",
    "Normal Score": "Normal Score",
    "Normal+Sensitive Score": "Normal+Sensitive Score",
    "Combination Score": "Combination Score",
    "Combination+Sensitive Score": "Combination+Sensitive Score",
    "Excessive Dryness score": "Excessive Dryness score",
    "Pregnancy Score": "Pregnancy Score",
    "Breastfeeling Score": "Breastfeeling Score",
}

FACE_LABEL_ALIASES = {
    "open pores": "Open Pores",
    "dark spots": "Dark Spots/Pigmentation",
    "pigmentation": "Dark Spots/Pigmentation",
    "uneven skin": "Uneven Skin Tone",
    "wrinkles": "Wrinkles/Fine lines",
    "redness": "Redness/Irritation",
    "barrier repair": "Barrier Repair",
}

NON_INGREDIENT_PHRASES = [
    "ingredient lists may change",
    "please refer to the product package",
    "refer the label",
    "ingredient listing",
    "leaping bunny certified",
    "peta and vegan certified",
    "provides everyday essentials",
    "what it smells like",
    "explore a wide range",
    "a key factor",
    "a prized ingredient",
    "as mentioned on the back",
    "lasting results",
    "natural ingredients",
    "abs plastic",
    "breathable elastic fabric",
    "breathable material",
    "bleached kraft pulp",
    "blotting paper",
]

NON_INGREDIENT_STARTS = [
    "it ",
    "this ",
    "that ",
    "what it ",
    "a key factor ",
    "a prized ingredient ",
    "acts as ",
    "act as ",
    "aids ",
    "aids in ",
    "helps ",
    "help ",
    "boosts ",
    "balances ",
    "brightens ",
    "cleans ",
    "cleanses ",
    "fights ",
    "maintaining ",
    "provides ",
    "reduces ",
    "soothes ",
    "hydrates ",
    "hydrate ",
    "known for ",
    "suitable for ",
    "treats ",
    "anti inflammatory properties ",
    "but also ",
]

NON_INGREDIENT_EXACT = {
    "ageing skin",
    "aging skin",
    "after bath body oil",
    "ingredient listing",
    "al oil",
    "amazonian rain forest ingredient",
    "aromatic",
    "broom",
    "but also",
    "cherry bomb flavor gloss",
}

KNOWN_COMPOUND_TOKENS = [
    "Papaya",
    "Sandalwood",
    "Saffron",
    "Rice",
    "Licorice",
    "Liquorice",
    "Neem",
    "Green Tea",
    "Coconut",
    "Walnut",
    "Nicotinamide",
    "Glycerin",
    "Glycerine",
    "Lemon Peel Extract",
    "Vitamin E",
    "Vitamin C",
    "Marigold Oil",
    "Wheat Germ Oil",
    "Kokum Butter",
    "Turmeric",
    "Aloe Vera",
    "Manjistha",
    "Ashwagandha",
    "Witch Hazel",
    "White Peony",
    "Eucalyptus Oil",
]

INGREDIENT_SOURCE_ALIASES = {
    "aloe vera": ["aloe vera", "aloe barbadensis"],
    "bha": ["salicylic acid", "salix alba", "willow bark", "beta hydroxy"],
    "salicylic acid": ["salicylic acid"],
    "willow bark extract": ["willow bark", "salix alba"],
    "glycolic acid": ["glycolic acid"],
    "lactic acid": ["lactic acid"],
    "kojic acid": ["kojic acid", "kojic dipalmitate"],
    "kojic dipalmitate": ["kojic acid", "kojic dipalmitate"],
    "green tea": ["green tea", "camellia sinensis"],
    "green tea extract": ["green tea", "camellia sinensis"],
    "fragrance": ["fragrance", "parfum", "perfume"],
    "lemon peel extract": ["lemon peel", "citrus limon", "citrus medica limonum"],
    "orange peel powder": ["orange peel", "citrus aurantium dulcis", "citrus sinensis"],
    "kaolin": ["kaolin"],
    "clay": ["kaolin", "bentonite", "multani", "fuller earth", "clay"],
    "lactobacillus ferment": ["lactobacillus ferment"],
    "probiotics ferments": ["lactobacillus ferment", "ferment", "bifida"],
    "probiotics": ["lactobacillus ferment", "ferment", "bifida"],
    "hyaluronic acid": ["hyaluronic acid", "sodium hyaluronate"],
    "niacinamide": ["niacinamide", "nicotinamide"],
}

TREATMENT_ACTIVE_TERMS = [
    "salicylic",
    "bha",
    "lactic acid",
    "glycolic",
    "mandelic",
    "kojic",
    "azelaic",
    "retinol",
    "retinal",
    "retinoid",
    "vitamin c",
    "ascorbic",
    "niacinamide",
    "willow bark",
]

PRESERVATIVE_OR_THRESHOLD_TERMS = [
    "phenoxyethanol",
    "ethylhexylglycerin",
    "sodium benzoate",
    "potassium sorbate",
    "benzyl alcohol",
    "edta",
    "disodium edta",
    "tetrasodium edta",
    "triethanolamine",
]

CLEANSER_SURFACTANT_TERMS = [
    "cocamidopropyl betaine",
    "sodium cocoyl",
    "sodium lauroyl",
    "sodium laureth",
    "sodium lauryl",
    "decyl glucoside",
    "lauryl glucoside",
    "coco glucoside",
    "caprylyl glucoside",
    "polysorbate",
    "peg 7 glyceryl cocoate",
    "disodium cocoamphodiacetate",
    "sodium methyl cocoyl taurate",
]

TREATMENT_CONCERN_COLUMNS = [
    "Acne",
    "Body Acne",
    "Open Pores",
    "Comedones",
    "Dark Spots/Pigmentation",
    "Uneven Skin Tone",
    "Melasma",
    "Dullness",
    "Tanning",
    "Wrinkles/Fine lines",
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split()).strip()


def norm_label(value: Any) -> str:
    text = clean_text(value).lower()
    text = text.replace("glycerine", "glycerin")
    text = text.replace("sulphate", "sulfate")
    text = text.replace("liquorice", "licorice")
    text = text.replace("co-enzyme", "coenzyme")
    text = re.sub(r"\bvit\s+", "vitamin ", text)
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split()).strip()


def norm_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", norm_label(value))


def without_percent(value: str) -> str:
    return clean_text(re.sub(r"\b\d+(?:\.\d+)?\s*%\s*", "", value))


def full_inci_from_retailer(row: dict[str, Any] | None) -> tuple[str, bool]:
    if row is None:
        return "", False
    ingredients = clean_text(row.get("ingredients"))
    if not ingredients:
        return "", False
    match = re.search(r"full\s+ingredient\s+list\s*:\s*(.+)", ingredients, flags=re.IGNORECASE)
    if match:
        return clean_text(match.group(1)), True
    return ingredients, False


def ingredient_presence_terms(label: str) -> list[str]:
    normalized = norm_label(without_percent(label))
    terms = [normalized]
    terms.extend(INGREDIENT_SOURCE_ALIASES.get(normalized, []))

    if "salicylic" in normalized or normalized == "bha":
        terms.extend(INGREDIENT_SOURCE_ALIASES["salicylic acid"])
    if "willow" in normalized or "salix" in normalized:
        terms.extend(INGREDIENT_SOURCE_ALIASES["willow bark extract"])
    if "glycolic" in normalized:
        terms.extend(INGREDIENT_SOURCE_ALIASES["glycolic acid"])
    if "lactic" in normalized:
        terms.extend(INGREDIENT_SOURCE_ALIASES["lactic acid"])
    if "kojic" in normalized:
        terms.extend(INGREDIENT_SOURCE_ALIASES["kojic acid"])
    if "green tea" in normalized:
        terms.extend(INGREDIENT_SOURCE_ALIASES["green tea"])
    if "aloe" in normalized:
        terms.extend(INGREDIENT_SOURCE_ALIASES["aloe vera"])
    if "probiotic" in normalized or "ferment" in normalized:
        terms.extend(INGREDIENT_SOURCE_ALIASES["probiotics"])
    if "fragrance" in normalized or "parfum" in normalized or "perfume" in normalized:
        terms.extend(INGREDIENT_SOURCE_ALIASES["fragrance"])

    output: list[str] = []
    for term in terms:
        cleaned = norm_label(term)
        if cleaned and cleaned not in output:
            output.append(cleaned)
    return output


def source_contains_ingredient(label: str, inci_text: str) -> bool:
    if not inci_text:
        return True
    label_norm = norm_label(without_percent(label))
    inci_norm = norm_label(inci_text)
    inci_key = norm_key(inci_text)

    if "fruit extract" in label_norm or "fruit complex" in label_norm:
        fruit_hits = [
            "fruit extract",
            "sugar cane",
            "orange fruit",
            "citrus aurantium",
            "vaccinium",
            "myrtillus",
            "acer saccharinum",
            "sugar maple",
        ]
        return sum(1 for term in fruit_hits if norm_label(term) in inci_norm) >= 2

    if "emollient cream base" in label_norm:
        emollient_hits = ["stearic acid", "glyceryl stearate", "caprylic capric triglyceride", "cetostearyl alcohol"]
        return any(norm_label(term) in inci_norm for term in emollient_hits)

    for term in ingredient_presence_terms(label):
        if len(term) < 3:
            continue
        if " " in term and term in inci_norm:
            return True
        if norm_key(term) and norm_key(term) in inci_key:
            return True
    return False


def source_validated_ingredients(
    labels: list[str],
    retailer_row: dict[str, Any] | None,
    group: str,
) -> tuple[list[str], list[str]]:
    inci_text, has_explicit_full_inci = full_inci_from_retailer(retailer_row)
    if not inci_text or not has_explicit_full_inci:
        return labels, []

    kept: list[str] = []
    flags: list[str] = []
    for label in labels:
        if source_contains_ingredient(label, inci_text):
            kept.append(label)
        else:
            flags.append(f"{group} ingredient removed: {label} was not found in retailer full INCI")
    return kept, flags


def split_full_inci_items(inci_text: str) -> list[str]:
    return [clean_text(part) for part in re.split(r"[,;]", inci_text) if clean_text(part)]


def ingredient_position(label: str, inci_items: list[str]) -> int | None:
    for index, item in enumerate(inci_items):
        if source_contains_ingredient(label, item):
            return index
    return None


def is_treatment_active(label: str) -> bool:
    normalized = norm_label(label)
    return any(term in normalized for term in TREATMENT_ACTIVE_TERMS)


def formula_quality_flags(
    product_type: str,
    primary_labels: list[str],
    secondary_labels: list[str],
    retailer_row: dict[str, Any] | None,
) -> list[str]:
    inci_text, has_explicit_full_inci = full_inci_from_retailer(retailer_row)
    if not inci_text or not has_explicit_full_inci:
        return []

    inci_items = split_full_inci_items(inci_text)
    if not inci_items:
        return []

    flags: list[str] = []
    preservative_index = None
    for index, item in enumerate(inci_items):
        if any(term in norm_label(item) for term in PRESERVATIVE_OR_THRESHOLD_TERMS):
            preservative_index = index
            break

    for label in primary_labels + secondary_labels:
        if not is_treatment_active(label):
            continue
        position = ingredient_position(label, inci_items)
        if position is None:
            continue
        late_by_position = position >= max(5, math.floor(len(inci_items) * 0.55))
        after_preservative = preservative_index is not None and position > preservative_index
        if late_by_position or after_preservative:
            flags.append(f"late or unknown-strength active: {label} appears low in retailer full INCI")

    if norm_label(product_type) == "cleanser":
        inci_norm = norm_label(inci_text)
        if not any(term in inci_norm for term in CLEANSER_SURFACTANT_TERMS):
            flags.append("cleanser full INCI has no clear primary surfactant")

    return flags


def apply_formula_quality_score_adjustments(
    scores: dict[str, int],
    product_type: str,
    flags: list[str],
) -> dict[str, int]:
    output = dict(scores)
    normalized_type = norm_label(product_type)
    has_late_active = any("late or unknown-strength active" in flag for flag in flags)
    has_unclear_cleanser_base = any("no clear primary surfactant" in flag for flag in flags)

    if normalized_type == "cleanser" and has_late_active:
        for column in TREATMENT_CONCERN_COLUMNS:
            if column in output:
                output[column] = min(output[column], 70 if column in {"Acne", "Body Acne", "Open Pores", "Comedones"} else 66)

    if normalized_type == "cleanser" and has_unclear_cleanser_base:
        for column in ["Oily Score", "Combination Score", "Normal Score"]:
            if column in output:
                output[column] = min(output[column], 76)
        for column in ["Oily+Sensitive Score", "Combination+Sensitive Score", "Normal+Sensitive Score", "Dry+Sensitive Score"]:
            if column in output:
                output[column] = min(output[column], 68)

    return output


def singular_key(key: str) -> str:
    if len(key) > 5 and key.endswith("ies"):
        return key[:-3] + "y"
    if len(key) > 4 and key.endswith("es"):
        return key[:-2]
    if len(key) > 4 and key.endswith("s"):
        return key[:-1]
    return key


def safe_float(value: Any) -> float | None:
    if value is None or clean_text(value) == "":
        return None
    try:
        numeric = float(str(value).replace("%", "").strip())
    except ValueError:
        return None
    if math.isnan(numeric):
        return None
    return numeric


def score_value(value: float) -> int:
    if value >= 0:
        return int(value + 0.5)
    return int(value - 0.5)


def average_with_hard_block(scores: list[float]) -> int | None:
    valid = [float(score) for score in scores if score is not None]
    if not valid:
        return None
    if any(score <= -100 for score in valid):
        return -100
    return score_value(sum(valid) / len(valid))


def weighted_average_with_hard_block(pairs: list[tuple[float, float]]) -> int | None:
    valid = [(float(score), float(weight)) for score, weight in pairs if score is not None and weight > 0]
    if not valid:
        return None
    if any(score <= -100 for score, _ in valid):
        return -100
    total_weight = sum(weight for _, weight in valid)
    if total_weight <= 0:
        return None
    return score_value(sum(score * weight for score, weight in valid) / total_weight)


def default_scores(age: int = 50, concern: int = 40, skin: int = 40, special: int = 50) -> dict[str, int]:
    scores: dict[str, int] = {}
    for col in AGE_COLUMNS:
        scores[col] = age
    for col in CONCERN_COLUMNS:
        scores[col] = concern
    for col in SKIN_COLUMNS:
        scores[col] = skin
    for col in SPECIAL_COLUMNS:
        scores[col] = special
    scores["None"] = 100
    return scores


def set_scores(scores: dict[str, int], columns: list[str], value: int) -> None:
    for col in columns:
        scores[col] = value


def recompute_unknown_skin_scores(scores: dict[str, int]) -> None:
    base = [scores.get(col, 40) for col in ["Oily Score", "Dry Score", "Normal Score", "Combination Score"]]
    sensitive = [
        scores.get(col, 40)
        for col in [
            "Oily+Sensitive Score",
            "Dry+Sensitive Score",
            "Normal+Sensitive Score",
            "Combination+Sensitive Score",
        ]
    ]
    scores["I dont know Score"] = average_with_hard_block(base) or 40
    scores["I dont know+Sensitive Score"] = average_with_hard_block(sensitive) or 40


def fallback_scores_for(label: str) -> tuple[str, dict[str, int], str]:
    normalized = norm_label(label)
    scores = default_scores()
    family = "Neutral functional ingredient"
    rationale = "Created because no reliable canonical match existed; conservative neutral/tolerated scores used."

    if any(token in normalized for token in ["fragrance", "parfum", "essential oil", "menthol"]):
        family = "Irritation-risk fragrant/cooling ingredient"
        scores = default_scores(age=40, concern=40, skin=40, special=50)
        set_scores(scores, ["Dryness", "Barrier Repair", "Redness/Irritation", "Dehydration"], -100)
        set_scores(scores, ["Dry Score", "Dry+Sensitive Score", "Oily+Sensitive Score", "Normal+Sensitive Score", "Combination+Sensitive Score"], -100)
        scores["Excessive Dryness score"] = 0
        rationale = "Fragrant/cooling ingredients are treated cautiously for dry, irritated, or sensitive profiles."
    elif any(token in normalized for token in ["alcohol", "denat"]):
        family = "Drying alcohol or uncertain alcohol"
        scores = default_scores(age=40, concern=40, skin=40, special=50)
        set_scores(scores, ["Dryness", "Barrier Repair", "Redness/Irritation", "Dehydration"], 0)
        set_scores(scores, ["Dry Score", "Dry+Sensitive Score", "Oily+Sensitive Score", "Normal+Sensitive Score", "Combination+Sensitive Score"], 0)
        scores["Excessive Dryness score"] = 0
        rationale = "Generic alcohol was not treated as a hard blocker, but was penalized for dryness and sensitivity."
    elif any(token in normalized for token in ["uv filter", "sunscreen filter", "spf", "sun filter"]):
        family = "Generic UV filter"
        scores = default_scores(age=100, concern=40, skin=90, special=50)
        set_scores(scores, ["Uneven Skin Tone", "Dark Spots/Pigmentation", "Melasma", "Tanning"], 90)
        scores["Tanning"] = 100
        scores["Excessive Dryness score"] = 50
        rationale = "Generic UV filters support photo-protection concerns; pregnancy/breastfeeding are neutral pending exact filter review."
    elif any(token in normalized for token in ["surfactant", "cleansing", "sles", "sls", "sulfate", "sulphate", "betaine", "soap"]):
        family = "Generic cleansing surfactant"
        scores = default_scores(age=100, concern=40, skin=40, special=50)
        set_scores(scores, ["Acne", "Body Acne", "Open Pores", "Comedones"], 90)
        set_scores(scores, ["Oily Score", "Combination Score"], 90)
        set_scores(scores, ["Dry Score", "Dry+Sensitive Score", "Oily+Sensitive Score", "Normal+Sensitive Score"], 40)
        scores["Excessive Dryness score"] = 0
        scores["Pregnancy Score"] = 100
        scores["Breastfeeling Score"] = 100
        rationale = "Generic cleansing agents fit oily/acne cleansing use but are cautious for dry or sensitive profiles."
    elif any(token in normalized for token in ["humectant", "hydrating", "hydration", "hyaluronic", "sorbitol", "glycerin", "glucose", "trehalose", "sugar"]):
        family = "Humectant or hydration support"
        scores = default_scores(age=100, concern=40, skin=90, special=100)
        set_scores(scores, ["Dryness", "Barrier Repair", "Redness/Irritation", "Dehydration", "Dullness"], 90)
        set_scores(scores, SKIN_COLUMNS, 90)
        scores["Dry Score"] = 100
        scores["Dry+Sensitive Score"] = 100
        rationale = "Hydration-focused fallback created from humectant behavior."
    elif any(
        token in normalized
        for token in [
            "emollient",
            "conditioning",
            "barrier repair",
            "petrolatum",
            "paraffin",
            "mineral oil",
            "dimethicone",
            "silicone",
            "lanolin",
            "wax",
            "butter",
            "malate",
            "polyisobutene",
            "polydecene",
            "oil",
            "fatty acid",
        ]
    ):
        family = "Emollient, occlusive, or barrier support"
        scores = default_scores(age=100, concern=40, skin=70, special=100)
        set_scores(scores, ["Dryness", "Barrier Repair", "Redness/Irritation", "Dehydration"], 90)
        set_scores(scores, ["Dry Score", "Dry+Sensitive Score", "Normal Score", "Normal+Sensitive Score"], 90)
        set_scores(scores, ["Oily Score", "Oily+Sensitive Score", "Combination Score"], 40)
        scores["Excessive Dryness score"] = 100
        rationale = "Emollient/occlusive fallback supports barrier and dryness but is conservative for oily or acne-prone profiles."
    elif any(
        token in normalized
        for token in [
            "brightening",
            "kojic",
            "arbutin",
            "licorice",
            "mulberry",
            "lemon",
            "orange peel",
            "orange powder",
            "orange fruit",
            "vitamin c",
            "pigment",
            "spot",
        ]
    ):
        family = "Brightening or pigmentation active"
        scores = default_scores(age=50, concern=40, skin=70, special=50)
        set_scores(scores, ["Uneven Skin Tone", "Dark Spots/Pigmentation", "Melasma", "Dullness", "Tanning"], 90)
        set_scores(scores, ["Oily Score", "Normal Score", "Combination Score"], 90)
        set_scores(scores, ["Dry Score", "Dry+Sensitive Score", "Oily+Sensitive Score", "Normal+Sensitive Score", "Combination+Sensitive Score"], 50)
        rationale = "Brightening fallback gives benefit to pigmentation-related concerns while staying cautious for teen and special-condition profiles."
    elif any(token in normalized for token in ["anti aging", "anti ageing", "wrinkle", "retinol", "retinal", "retinoid", "peptide", "collagen", "coenzyme"]):
        family = "Anti-aging support active"
        scores = default_scores(age=50, concern=40, skin=70, special=50)
        set_scores(scores, ["Wrinkles/Fine lines", "Dullness"], 90)
        set_scores(scores, ["Dryness", "Dehydration"], 70)
        set_scores(scores, ["Oily Score", "Normal Score", "Combination Score"], 90)
        set_scores(scores, ["Dry Score", "Dry+Sensitive Score", "Oily+Sensitive Score", "Normal+Sensitive Score", "Combination+Sensitive Score"], 50)
        rationale = "Generic anti-aging active fallback is useful for wrinkles but neutral/cautious for teen, dry-sensitive, pregnancy, and breastfeeding."
    elif any(token in normalized for token in ["barrier", "ceramide", "cholesterol", "fatty acid"]):
        family = "Barrier repair active"
        scores = default_scores(age=100, concern=40, skin=90, special=100)
        set_scores(scores, ["Dryness", "Barrier Repair", "Redness/Irritation", "Dehydration"], 90)
        set_scores(scores, SKIN_COLUMNS, 90)
        scores["Dry Score"] = 100
        scores["Dry+Sensitive Score"] = 100
        rationale = "Barrier-repair fallback created from the ingredient label."
    elif any(token in normalized for token in ["pore", "blemish", "acne", "salicylic", "tea tree", "zinc", "clay", "charcoal"]):
        family = "Acne, pore, or blemish active"
        scores = default_scores(age=50, concern=40, skin=40, special=50)
        set_scores(scores, ["Acne", "Body Acne", "Open Pores", "Comedones"], 90)
        set_scores(scores, ["Oily Score", "Combination Score"], 90)
        set_scores(scores, ["Dry Score", "Dry+Sensitive Score", "Oily+Sensitive Score", "Normal+Sensitive Score", "Combination+Sensitive Score"], 0)
        scores["Excessive Dryness score"] = 0
        rationale = "Acne/pore active fallback prioritizes oily/acne fit and penalizes dry-sensitive profiles."
    elif any(token in normalized for token in ["exfoliant", "exfoliating", "scrub", "aha", "bha", "lactic", "glycolic", "mandelic"]):
        family = "Exfoliating active"
        scores = default_scores(age=50, concern=40, skin=40, special=50)
        set_scores(scores, ["Acne", "Open Pores", "Comedones", "Uneven Skin Tone", "Dullness", "Tanning"], 90)
        set_scores(scores, ["Dryness", "Barrier Repair", "Redness/Irritation", "Dehydration"], 0)
        set_scores(scores, ["Oily Score", "Combination Score"], 90)
        set_scores(scores, ["Dry Score", "Dry+Sensitive Score", "Oily+Sensitive Score", "Normal+Sensitive Score", "Combination+Sensitive Score"], 0)
        scores["Excessive Dryness score"] = 0
        rationale = "Exfoliating fallback gives concern benefit but is cautious for compromised barrier or sensitive profiles."
    elif any(token in normalized for token in ["probiotic", "ferment", "mucin", "snail", "heartleaf", "cica", "centella", "green tea", "cucumber", "aloe", "oat", "chamomile"]):
        family = "Soothing or microbiome-support ingredient"
        scores = default_scores(age=100, concern=40, skin=90, special=100)
        set_scores(scores, ["Barrier Repair", "Redness/Irritation", "Dehydration", "Dullness"], 90)
        set_scores(scores, SKIN_COLUMNS, 90)
        rationale = "Soothing fallback supports barrier, redness, hydration, and broad skin-type fit."
    elif any(token in normalized for token in ["antioxidant", "vitamin e", "tocopherol", "acai", "berry"]):
        family = "Antioxidant support"
        scores = default_scores(age=100, concern=40, skin=90, special=100)
        set_scores(scores, ["Uneven Skin Tone", "Dark Spots/Pigmentation", "Wrinkles/Fine lines", "Dullness", "Tanning"], 90)
        rationale = "Antioxidant fallback supports dullness, photo-stress, tone, and aging-related concerns."

    recompute_unknown_skin_scores(scores)
    return family, scores, rationale


def is_non_ingredient_text(label: str) -> bool:
    normalized = norm_label(label)
    if not normalized:
        return True
    if normalized in NON_INGREDIENT_EXACT:
        return True
    if len(normalized) > 120:
        return True
    if any(normalized.startswith(prefix) for prefix in NON_INGREDIENT_STARTS):
        return True
    if len(normalized.split()) >= 8 and any(
        token in normalized
        for token in [" helps ", " hydrates ", " cleans ", " soothes ", " fights ", " reduces ", " improves ", " repairs "]
    ):
        return True
    return any(phrase in normalized for phrase in NON_INGREDIENT_PHRASES)


def read_ingredient_scores() -> tuple[dict[str, dict[str, Any]], dict[str, str], dict[str, str], list[dict[str, Any]]]:
    workbook = openpyxl.load_workbook(INGREDIENT_SCORES, read_only=True, data_only=True)
    sheet = workbook["Ingredient Scores"]
    raw_headers = [clean_text(value) for value in next(sheet.iter_rows(values_only=True))]
    canonical_rows: dict[str, dict[str, Any]] = {}
    alias_map: dict[str, str] = {}
    alias_labels: dict[str, str] = {}
    original_rows: list[dict[str, Any]] = []

    def add_alias(alias: str, canonical: str) -> None:
        key = norm_key(alias)
        if key:
            alias_map.setdefault(key, canonical)
            alias_labels.setdefault(key, clean_text(alias))
            alias_map.setdefault(singular_key(key), canonical)
            alias_labels.setdefault(singular_key(key), clean_text(alias))

    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(raw_headers, raw_row))
        canonical = clean_text(row.get("Canonical Ingredient"))
        if not canonical:
            continue
        scores: dict[str, int] = {}
        for source_header, target_header in INGREDIENT_SCORE_HEADER_MAP.items():
            value = safe_float(row.get(source_header))
            scores[target_header] = score_value(value) if value is not None else 40
        scores["None"] = 100
        recompute_unknown_skin_scores(scores)
        canonical_rows[canonical] = {
            "canonical_ingredient": canonical,
            "source": "Roopsee ingredient score sheet",
            "skintype_match": clean_text(row.get("Skintype match")),
            "concerns": clean_text(row.get("Concerns")),
            "scores": scores,
            "fallback_family": "",
            "fallback_rationale": "",
        }
        original_rows.append(canonical_rows[canonical])

        add_alias(canonical, canonical)
        no_percent = without_percent(canonical)
        add_alias(no_percent, canonical)
        no_parentheses = clean_text(re.sub(r"\([^)]*\)", "", no_percent))
        add_alias(no_parentheses, canonical)
        for part in re.split(r"\s*/\s*|\s+\+\s+|\s+and\s+|\s+or\s+", no_percent, flags=re.I):
            add_alias(part, canonical)
        for parenthetical in re.findall(r"\(([^)]*)\)", canonical):
            add_alias(parenthetical, canonical)

    return canonical_rows, alias_map, alias_labels, original_rows


def build_manual_aliases(canonical_rows: dict[str, dict[str, Any]]) -> dict[str, str]:
    candidates = {
        "jojoba oil": ["Jojoba Oils", "Natural Golden Jojoba Oil (Simmondsia Chinensis)"],
        "green tea": ["Green Tea Extract"],
        "licorice": ["Licorice Extract"],
        "tea tree": ["Tea Tree Extract"],
        "kojic acid": ["Kojic Dipalmitate"],
        "ceramide complex": ["Ceramides"],
        "mucin": ["Snail Mucin"],
        "glycerine": ["Glycerin", "Plant-Based Glycerine"],
        "sunflower oil": ["Sunflower Seed Oil", "Sunflower"],
        "heartleaf": ["Heartleaf Extract"],
        "amino acids": ["Amino Acid"],
        "coenzyme q10": ["Ubiquinone", "Coenzyme Q10"],
        "sodium laureth sulphate": ["Sodium Laureth Sulfate / SLES"],
        "sodium laureth sulfate": ["Sodium Laureth Sulfate / SLES"],
        "sodium lauryl sulphate": ["Sodium Lauryl Sulfate / SLS / SDS"],
        "sodium lauryl sulfate": ["Sodium Lauryl Sulfate / SLS / SDS"],
        "cocoamidopropyl betaine": ["Cocamidopropyl Betaine"],
        "cocamidopropyl betaine": ["Cocamidopropyl Betaine"],
        "lavender oil": ["Lavender Essential Oil"],
        "rosemary oil": ["Rosemary Essential Oil"],
        "jasmine": ["Jasmine Essential Oil"],
        "lemon": ["Lemon Extract", "Lemon Peel Extract"],
        "cucumber": ["Cucumber Extract"],
        "collagen": ["Collagen"],
        "marigold": ["Marigold Extract"],
        "amla": ["Amla Extract"],
        "gooseberry": ["Amla Extract"],
        "witch hazel": ["Witch Hazel", "Hamamelis Virginiana (Witch Hazel)", "American Witch Hazel Extract"],
        "hyssop extract": ["Hyssop Extract"],
        "sage": ["Sage Extract"],
        "moringa": ["Moringa Extract", "Moringa"],
        "sigru extract": ["Moringa Extract", "Moringa"],
        "rumex occidentalis": ["Rumex Crispus Root Extract"],
        "rumex": ["Rumex Crispus Root Extract"],
        "nariyal": ["Coconut Oil", "Virgin Coconut Oil"],
        "jaitun": ["Olive Oil", "Olive"],
        "mulethi": ["Mulethi Powder", "Licorice Extract"],
        "mulethi root": ["Mulethi Powder", "Licorice Extract"],
        "multani mitti": ["Multani Mitti", "Multani Mitti Powder"],
        "fullers earth": ["Multani Mitti", "Multani Mitti Powder"],
        "wheat germ oil": ["Wheat Germ Oil"],
        "almond oil": ["Almond Oil", "Sweet Almond Oil"],
        "rose": ["Rose Extract", "Rose Water"],
        "rosemary": ["Rosemary", "Rosemary Leaf Extract"],
        "apple cider vinegar": ["Apple Cider Vinegar"],
        "haldi": ["Turmeric Extract", "Turmeric"],
        "turmeric": ["Turmeric Extract", "Turmeric"],
        "chandan": ["Sandalwood Extract", "Sandalwood"],
        "chandana": ["Sandalwood Extract", "Sandalwood"],
        "safed chandan": ["Sandalwood Extract", "Sandalwood"],
        "rakt chandan": ["Red Sandalwood Extract", "Sandalwood Extract", "Sandalwood"],
        "kumari": ["Aloe Vera", "Aloe Vera Extract"],
        "kumari leaf pulp": ["Aloe Vera", "Aloe Vera Extract"],
        "ghritkumari leaf": ["Aloe Vera", "Aloe Vera Extract"],
        "cermaide": ["Ceramides"],
        "cera hyamino": ["Ceramides"],
        "camomile": ["Chamomile Extract"],
        "chamomile": ["Chamomile Extract"],
        "manjistha": ["Manjistha", "Manjistha (Indian Madder)"],
        "white peony": ["White peony", "Peony Extract"],
        "peony": ["Peony Extract", "Peony"],
        "eucalyptus oil": ["Eucalyptus essential oil", "Eucalyptus Globulus Leaf Oil"],
        "eucalyptus": ["Eucalyptus Globulus Leaf Oil", "Eucalyptus essential oil"],
        "bearberry fruit juice": ["Bearberry Extract", "Bearberry Leaf Extract"],
        "berry": ["Berry Extract"],
        "berries": ["Berry Extract"],
        "cherry": ["Cherry Extract"],
        "cherry fruit juice": ["Cherry Extract"],
    }
    manual: dict[str, str] = {}
    for label, preferred in candidates.items():
        for candidate in preferred:
            if candidate in canonical_rows:
                manual[norm_key(label)] = candidate
                break
    return manual


def candidate_match(
    label: str,
    alias_map: dict[str, str],
    alias_labels: dict[str, str],
    canonical_rows: dict[str, dict[str, Any]],
) -> tuple[str, str, float] | None:
    key = norm_key(without_percent(label))
    if not key:
        return None
    if key in alias_map:
        return alias_map[key], "Exact or alias", 1.0
    singular = singular_key(key)
    if singular in alias_map:
        return alias_map[singular], "Singular/plural normalized", 0.98

    normalized = norm_label(without_percent(label))
    tokens = set(normalized.split())
    if len(tokens) >= 2:
        best: tuple[str, float, str] | None = None
        for alias_key, canonical in alias_map.items():
            alias_label = norm_label(alias_labels.get(alias_key) or canonical)
            alias_tokens = set(alias_label.split())
            if not alias_tokens:
                continue
            overlap = len(tokens & alias_tokens) / max(1, len(tokens | alias_tokens))
            ratio = difflib.SequenceMatcher(None, key, alias_key).ratio()
            if (tokens <= alias_tokens or alias_tokens <= tokens) and overlap >= 0.5:
                score = max(0.86, overlap)
                if best is None or score > best[1]:
                    best = (canonical, score, "Token subset relation")
            elif ratio >= 0.9:
                if best is None or ratio > best[1]:
                    best = (canonical, ratio, "Fuzzy normalized relation")
        if best:
            return best[0], best[2], best[1]

    return None


def split_ingredients(value: str) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    text = re.sub(r"(?<=\d),(?=\d{3}\b)", "", text)
    text = re.sub(r"\bIngredients?\s*:\s*", "", text, flags=re.I)
    text = re.sub(r"\bKey Ingredients?\s*:\s*", "", text, flags=re.I)
    text = text.replace("•", ",").replace("|", ",")
    parts = re.split(r",|;|\n|\s+\+\s+", text)
    output: list[str] = []
    for raw_part in parts:
        part = without_percent(raw_part)
        part = re.sub(r"\([^)]*(?:refer|please|package|change|label)[^)]*\)", "", part, flags=re.I)
        part = re.sub(r"\b(?:usda\s+organic\s+certified|certified\s+organic|organic-certified|organic)\b[-\s]*", "", part, flags=re.I)
        part = re.sub(r"\b(?:also\s+known\s+as|known\s+as)\b", "", part, flags=re.I)
        part = re.sub(r"^\s*(?:and|with)\s+", "", part, flags=re.I)
        if "." in part:
            after_period = clean_text(part.rsplit(".", 1)[1])
            if after_period and len(after_period) <= 80:
                part = after_period
        if ":" in part:
            before_colon = clean_text(part.split(":", 1)[0])
            if "." in before_colon:
                after_period = clean_text(before_colon.rsplit(".", 1)[1])
                if after_period and len(after_period) <= 80:
                    before_colon = after_period
            if before_colon and len(before_colon) <= 80:
                part = before_colon
        part = re.sub(r"\(\s*\d+\s*ppm\s*\)", "", part, flags=re.I)
        part = re.sub(r"\b\d+\s*ppm\b", "", part, flags=re.I)
        part = re.sub(
            r"\b(?:helps?|aids?|acts?|known|rich|due|cleans(?:es)?|hydrates?|soothes?|brightens?|fights?|reduces?|boosts?|balances?|maintains?|improves?|repairs?)\b.*$",
            "",
            part,
            flags=re.I,
        )
        part = re.sub(r"\bis\s+(?:ideal|the|a)\b.*$", "", part, flags=re.I)
        part = clean_text(part.strip(" -*_./"))
        if not part:
            continue
        if norm_label(part) in {"none", "na", "n a", "null", "refer label"}:
            continue
        compound = split_known_compound(part)
        if compound:
            output.extend(compound)
            continue
        output.append(part)
    return output


def split_known_compound(part: str) -> list[str]:
    if len(part) < 50:
        return []
    compact = norm_key(part)
    found: list[str] = []
    for token in KNOWN_COMPOUND_TOKENS:
        if norm_key(token) in compact and token not in found:
            found.append(token)
    return found if len(found) >= 2 else []


def repair_ingredient_columns(primary: str, secondary: str) -> tuple[str, str]:
    primary_clean = clean_text(primary)
    secondary_clean = clean_text(secondary)
    broken_quantity = bool(re.search(r"\(\s*\d+\s*$", primary_clean)) and bool(re.match(r"^\s*\d", secondary_clean))
    if broken_quantity:
        return f"{primary_clean}{secondary_clean}", ""
    return primary_clean, secondary_clean


def product_type_from(name: str, retailer_row: dict[str, Any] | None) -> str:
    name_text = norm_label(name)
    if any(token in name_text for token in ["sunscreen", "sun screen", "spf", "sunblock", "sun block"]):
        return "Sunscreen"
    if any(token in name_text for token in ["mask", "masque", "pack", "ubtan", "face powder", "skin powder"]):
        return "Mask"
    if any(token in name_text for token in ["face wash", "body wash", "cleanser", "cleansing foam", "cleansing gel", "wash gel"]):
        return "Cleanser"
    if any(token in name_text for token in ["body acne spray", "acne spray", "salyzap", "treatment spray"]):
        return "Toner"
    if "toner" in name_text or "toning" in name_text or "facial mist" in name_text:
        return "Toner"
    if any(token in name_text for token in ["serum", "ampoule", "booster", "concentrate"]):
        return "Serum"
    if any(token in name_text for token in ["moisturizer", "moisturiser", "cream", "lotion", "body butter", "balm", "gel cream"]):
        return "Moisturizer"

    text_parts = [name]
    if retailer_row:
        text_parts.extend(
            [
                retailer_row.get("product_attributes", ""),
                retailer_row.get("categories", ""),
                retailer_row.get("source_categories", ""),
            ]
        )
    text = norm_label(" ".join(clean_text(part) for part in text_parts))
    if any(token in text for token in ["sunscreen", "sun screen", "spf", "sunblock", "sun block"]):
        return "Sunscreen"
    if any(token in text for token in ["face wash", "body wash", "cleanser", "cleansing foam", "cleansing gel", "wash gel"]):
        return "Cleanser"
    if "toner" in text or "toning" in text or "facial mist" in text:
        return "Toner"
    if any(
        token in text
        for token in [
            "mask",
            "masque",
            "pack",
            "ubtan",
            "face powder",
            "skin powder",
            "powder for face",
            "powder for skin",
            "orange peel powder",
            "orange powder",
            "clay powder",
            "kaolin clay powder",
            "charcoal powder",
            "neem leaf",
            "turmeric powder",
            "licorice powder",
            "dried orange peel",
            "multani",
            "mitti",
            "chandan powder",
        ]
    ):
        return "Mask"
    if any(token in text for token in ["serum", "ampoule", "booster", "concentrate"]):
        return "Serum"
    if any(token in text for token in ["scrub", "exfoliator", "peel"]):
        return "Mask"
    if any(token in text for token in ["moisturizer", "moisturiser", "cream", "lotion", "body butter", "balm", "gel cream"]):
        return "Moisturizer"
    return "Other"


def product_category_from(name: str, product_type: str, retailer_row: dict[str, Any] | None) -> str:
    text = norm_label(
        " ".join(
            [
                name,
                retailer_row.get("categories", "") if retailer_row else "",
                retailer_row.get("source_categories", "") if retailer_row else "",
                retailer_row.get("product_attributes", "") if retailer_row else "",
            ]
        )
    )
    if "lip" in text:
        return "Lips"
    if "eye" in text or "under eye" in text or "dark circle" in text:
        return "Eye"
    body_terms = [
        "body wash",
        "bodywash",
        "body lotion",
        "body cream",
        "body butter",
        "body sunscreen",
        "body spray",
        "body acne",
        "back acne",
        "hand cream",
        "foot cream",
        "hand & foot",
        "hand and foot",
        "elbow",
        "knee",
        "bum",
        "buttock",
        "inner thigh",
    ]
    face_terms = [
        "face wash",
        "facewash",
        "face cleanser",
        "face serum",
        "face cream",
        "face moisturizer",
        "face moisturiser",
        "face mask",
        "face pack",
        "face powder",
        "skin powder",
        "facial",
        "toner",
        "sunscreen",
        "sun cream",
        "spf",
    ]
    has_body = any(term in text for term in body_terms) or product_type == "Body Wash"
    has_face = any(term in text for term in face_terms) or product_type in {"Serum", "Toner", "Sunscreen", "Mask"}
    if has_body and has_face:
        return "Face & Body"
    if has_body:
        return "Body"
    return "Face"


def reviewed_product_ingredient_override(
    product_name: str,
    product_type: str,
    primary_labels: list[str],
    secondary_labels: list[str],
) -> tuple[str, list[str], list[str]]:
    name = norm_label(product_name)
    if "be neude instant detanning and brightening face mask" in name:
        return (
            "Mask",
            ["Kaolin", "Milk Cream", "Isopropyl Myristate", "Fruit Extract Complex"],
            [
                "Lactic Acid",
                "Kojic Acid",
                "Lactobacillus Ferment",
                "Aloe Vera",
                "Willow Bark Extract",
                "Lemon Peel Extract",
                "Fragrance",
            ],
        )
    return product_type, primary_labels, secondary_labels


def product_type_rule(product_type: str) -> dict[str, Any]:
    normalized = norm_label(product_type)
    if normalized == "cleanser":
        return {"skin_type": True, "concern": "Yes except wrinkles", "primary_weight": 0.5, "secondary_weight": 0.5}
    if normalized == "toner":
        return {"skin_type": True, "concern": True, "primary_weight": 0.5, "secondary_weight": 0.5}
    if normalized == "serum":
        return {"skin_type": True, "concern": True, "primary_weight": 0.8, "secondary_weight": 0.2}
    if normalized in {"moisturizer", "moisturiser", "sunscreen"}:
        return {"skin_type": True, "concern": False, "primary_weight": 0.5, "secondary_weight": 0.5}
    if normalized == "mask":
        return {"skin_type": True, "concern": True, "primary_weight": 0.5, "secondary_weight": 0.5}
    return {"skin_type": True, "concern": True, "primary_weight": 0.5, "secondary_weight": 0.5}


def read_useful_products() -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(USEFUL_PRODUCTS, read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        name = clean_text(row[0] if len(row) > 0 else "")
        if not name:
            continue
        primary = clean_text(row[1] if len(row) > 1 else "")
        secondary = clean_text(row[2] if len(row) > 2 else "")
        primary, secondary = repair_ingredient_columns(primary, secondary)
        rows.append(
            {
                "source_row": row_number,
                "product_name": name,
                "product_name_key": norm_key(name),
                "primary_ingredients_text": primary,
                "secondary_ingredients_text": secondary,
                "primary_ingredients": split_ingredients(primary),
                "secondary_ingredients": split_ingredients(secondary),
            }
        )
    return rows


def read_retailer_rows() -> dict[str, list[dict[str, Any]]]:
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    with open_text(RETAILER_PRODUCTS) as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            key = norm_key(row.get("product_name"))
            if key:
                by_name[key].append(row)
    return by_name


def best_retailer_row(rows: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not rows:
        return None

    def sort_key(row: dict[str, Any]) -> tuple[int, float, str]:
        in_stock = 0 if clean_text(row.get("in_stock")).lower() == "true" else 1
        price = safe_float(row.get("selling_price")) or safe_float(row.get("mrp")) or 10**9
        return in_stock, price, clean_text(row.get("id"))

    return sorted(rows, key=sort_key)[0]


def retailer_value(row: dict[str, Any] | None, key: str, max_len: int = 3000) -> str:
    if row is None:
        return ""
    value = clean_text(row.get(key))
    if len(value) > max_len:
        return value[: max_len - 20] + " ...[truncated]"
    return value


def group_score(matches: list[dict[str, Any]], column: str) -> int | None:
    scores: list[float] = []
    for match in matches:
        if match.get("used_for_scoring") and match.get("scores"):
            score = match["scores"].get(column)
            if score is not None:
                scores.append(float(score))
    return average_with_hard_block(scores)


def product_scores(primary_matches: list[dict[str, Any]], secondary_matches: list[dict[str, Any]], product_type: str) -> dict[str, int]:
    rule = product_type_rule(product_type)
    primary_weight = float(rule["primary_weight"])
    secondary_weight = float(rule["secondary_weight"])
    if primary_matches and not secondary_matches:
        primary_weight, secondary_weight = 1.0, 0.0
    elif secondary_matches and not primary_matches:
        primary_weight, secondary_weight = 0.0, 1.0

    output: dict[str, int] = {}
    for column in SCORE_COLUMNS:
        primary_score = group_score(primary_matches, column)
        secondary_score = group_score(secondary_matches, column)
        score = weighted_average_with_hard_block(
            [
                (primary_score, primary_weight if primary_score is not None else 0),
                (secondary_score, secondary_weight if secondary_score is not None else 0),
            ]
        )
        if score is None:
            if column == "None":
                score = 100
            elif column in AGE_COLUMNS or column in SPECIAL_COLUMNS:
                score = 50
            else:
                score = 40
        output[column] = score
    output["None"] = 100
    recompute_unknown_skin_scores(output)
    return output


def make_matcher(canonical_rows: dict[str, dict[str, Any]], alias_map: dict[str, str], alias_labels: dict[str, str]):
    manual_aliases = build_manual_aliases(canonical_rows)
    fallback_rows: dict[str, dict[str, Any]] = {}
    match_cache: dict[str, dict[str, Any]] = {}

    def match(label: str) -> dict[str, Any]:
        cleaned = clean_text(label)
        cache_key = norm_key(cleaned)
        if cache_key in match_cache:
            cached = dict(match_cache[cache_key])
            cached["input_ingredient"] = cleaned
            return cached

        if not cleaned or is_non_ingredient_text(cleaned):
            result = {
                "input_ingredient": cleaned,
                "canonical_ingredient": "",
                "match_method": "Skipped non-ingredient text",
                "confidence": 0.0,
                "used_for_scoring": False,
                "scores": {},
                "fallback_family": "",
                "fallback_rationale": "Text looked like a disclaimer, marketing phrase, or unparseable non-ingredient.",
            }
            match_cache[cache_key] = result
            return dict(result)

        normalized_key = norm_key(without_percent(cleaned))
        if normalized_key in manual_aliases:
            canonical = manual_aliases[normalized_key]
            row = canonical_rows[canonical]
            result = {
                "input_ingredient": cleaned,
                "canonical_ingredient": canonical,
                "match_method": "Curated related match",
                "confidence": 0.95,
                "used_for_scoring": True,
                "scores": row["scores"],
                "fallback_family": row.get("fallback_family", ""),
                "fallback_rationale": "Matched through curated normalization for a common Roopsee ingredient label.",
            }
            match_cache[cache_key] = result
            return dict(result)

        matched = candidate_match(cleaned, alias_map, alias_labels, canonical_rows)
        if matched:
            canonical, method, confidence = matched
            row = canonical_rows[canonical]
            result = {
                "input_ingredient": cleaned,
                "canonical_ingredient": canonical,
                "match_method": method,
                "confidence": round(confidence, 3),
                "used_for_scoring": True,
                "scores": row["scores"],
                "fallback_family": row.get("fallback_family", ""),
                "fallback_rationale": "Matched to existing Roopsee ingredient score after normalization.",
            }
            match_cache[cache_key] = result
            return dict(result)

        family, scores, rationale = fallback_scores_for(cleaned)
        canonical = cleaned
        fallback_rows.setdefault(
            canonical,
            {
                "canonical_ingredient": canonical,
                "source": "Created fallback",
                "skintype_match": family,
                "concerns": "",
                "scores": scores,
                "fallback_family": family,
                "fallback_rationale": rationale,
            },
        )
        result = {
            "input_ingredient": cleaned,
            "canonical_ingredient": canonical,
            "match_method": "Created fallback",
            "confidence": 0.55,
            "used_for_scoring": True,
            "scores": scores,
            "fallback_family": family,
            "fallback_rationale": rationale,
        }
        match_cache[cache_key] = result
        return dict(result)

    return match, fallback_rows, match_cache


def summarize_match_status(matches: list[dict[str, Any]]) -> str:
    used = [match for match in matches if match.get("used_for_scoring")]
    if not used:
        return "Needs review - no scoreable ingredients"
    methods = {match["match_method"] for match in used}
    if methods <= {"Exact or alias", "Singular/plural normalized"}:
        return "Matched to ingredient score sheet"
    if "Created fallback" in methods:
        return "Mixed with created fallback"
    return "Matched with normalization"


def ingredient_summary(matches: list[dict[str, Any]]) -> str:
    parts = []
    for match in matches:
        if not match.get("input_ingredient"):
            continue
        canonical = match.get("canonical_ingredient") or "not used"
        parts.append(f"{match['input_ingredient']} -> {canonical} ({match['match_method']})")
    value = "; ".join(parts)
    return value[:3000] + " ...[truncated]" if len(value) > 3000 else value


def build_rows() -> dict[str, Any]:
    canonical_rows, alias_map, alias_labels, original_ingredient_rows = read_ingredient_scores()
    useful_products = read_useful_products()
    retailer_by_name = read_retailer_rows()
    match, fallback_rows, match_cache = make_matcher(canonical_rows, alias_map, alias_labels)

    unique_headers = [
        "auto_product_uid",
        "source_row",
        "product_name",
        "matched_retailer_sku_count",
        "site",
        "parent_product_id",
        "product_id",
        "sku",
        "brand",
        "variant",
        "category",
        "product_type",
        "mrp",
        "selling_price",
        "rating",
        "rating_count",
        "review_count",
        "in_stock",
        "product_url",
        "image_url",
        "primary_ingredients",
        "secondary_ingredients",
        "matched_primary_ingredients",
        "matched_secondary_ingredients",
        "ingredient_match_status",
        "created_fallback_ingredient_count",
        "needs_review_ingredient_count",
        "score_basis",
        "source_validation_flags",
        "formula_quality_flags",
        "retailer_full_ingredients",
        "retailer_how_to_use",
        "retailer_product_attributes",
    ] + SCORE_COLUMNS

    sku_headers = [
        "auto_product_uid",
        "source_row",
        "retailer_row_id",
        "site",
        "parent_product_id",
        "product_id",
        "sku",
        "brand",
        "product_name",
        "variant",
        "category",
        "product_type",
        "mrp",
        "selling_price",
        "discount_pct",
        "rating",
        "rating_count",
        "review_count",
        "in_stock",
        "product_url",
        "image_url",
        "primary_ingredients",
        "secondary_ingredients",
        "matched_primary_ingredients",
        "matched_secondary_ingredients",
        "ingredient_match_status",
        "created_fallback_ingredient_count",
        "needs_review_ingredient_count",
        "score_basis",
        "source_validation_flags",
        "formula_quality_flags",
        "retailer_key_ingredients",
        "retailer_full_ingredients",
        "retailer_how_to_use",
        "retailer_product_attributes",
    ] + SCORE_COLUMNS

    unique_rows: list[list[Any]] = []
    sku_rows: list[list[Any]] = []
    unmatched_products: list[list[Any]] = []
    match_audit_seen: dict[tuple[str, str], dict[str, Any]] = {}
    type_counter: Counter[str] = Counter()
    status_counter: Counter[str] = Counter()

    for product_index, product in enumerate(useful_products, start=1):
        retailer_rows = retailer_by_name.get(product["product_name_key"], [])
        representative = best_retailer_row(retailer_rows)
        product_type = product_type_from(product["product_name"], representative)
        product_type, raw_primary_labels, raw_secondary_labels = reviewed_product_ingredient_override(
            product["product_name"],
            product_type,
            product["primary_ingredients"],
            product["secondary_ingredients"],
        )
        category = product_category_from(product["product_name"], product_type, representative)
        type_counter[product_type] += 1

        primary_labels, primary_source_flags = source_validated_ingredients(raw_primary_labels, representative, "Primary")
        secondary_labels, secondary_source_flags = source_validated_ingredients(raw_secondary_labels, representative, "Secondary")
        source_validation_flags = primary_source_flags + secondary_source_flags
        quality_flags = formula_quality_flags(product_type, primary_labels, secondary_labels, representative)
        primary_text = ", ".join(primary_labels)
        secondary_text = ", ".join(secondary_labels)

        primary_matches = [match(label) for label in primary_labels]
        secondary_matches = [match(label) for label in secondary_labels]
        all_matches = primary_matches + secondary_matches
        scores = product_scores(primary_matches, secondary_matches, product_type)
        scores = apply_formula_quality_score_adjustments(scores, product_type, quality_flags)
        status = summarize_match_status(all_matches)
        if source_validation_flags:
            status = "Matched with source INCI correction"
        elif quality_flags:
            status = "Matched with formula quality cautions"
        status_counter[status] += 1
        created_count = sum(1 for item in all_matches if item.get("match_method") == "Created fallback")
        needs_review_count = sum(1 for item in all_matches if not item.get("used_for_scoring"))
        auto_uid_base = f"AUTO-{product_index:05d}"
        score_basis = (
            "Primary/secondary ingredient scores combined with hard-blocker handling; "
            f"product type rule: {product_type_rule(product_type)}"
        )
        if source_validation_flags or quality_flags:
            score_basis += f"; source validation: {'; '.join(source_validation_flags + quality_flags)}"

        for group, matches in [("Primary", primary_matches), ("Secondary", secondary_matches)]:
            for item in matches:
                key = (item.get("input_ingredient", ""), group)
                if key not in match_audit_seen:
                    match_audit_seen[key] = {
                        "input_ingredient": item.get("input_ingredient", ""),
                        "ingredient_group": group,
                        "canonical_ingredient": item.get("canonical_ingredient", ""),
                        "match_method": item.get("match_method", ""),
                        "confidence": item.get("confidence", 0),
                        "used_for_scoring": item.get("used_for_scoring", False),
                        "fallback_family": item.get("fallback_family", ""),
                        "fallback_rationale": item.get("fallback_rationale", ""),
                        "product_examples": [],
                    }
                examples = match_audit_seen[key]["product_examples"]
                if len(examples) < 5:
                    examples.append(product["product_name"])

        base_unique = [
            auto_uid_base,
            product["source_row"],
            product["product_name"],
            len(retailer_rows),
            retailer_value(representative, "site"),
            retailer_value(representative, "parent_product_id"),
            retailer_value(representative, "product_id"),
            retailer_value(representative, "sku"),
            retailer_value(representative, "brand"),
            retailer_value(representative, "variant"),
            category,
            product_type,
            safe_float(retailer_value(representative, "mrp")) if representative else "",
            safe_float(retailer_value(representative, "selling_price")) if representative else "",
            safe_float(retailer_value(representative, "rating")) if representative else "",
            safe_float(retailer_value(representative, "rating_count")) if representative else "",
            safe_float(retailer_value(representative, "review_count")) if representative else "",
            retailer_value(representative, "in_stock"),
            retailer_value(representative, "product_url"),
            retailer_value(representative, "image_url"),
            primary_text,
            secondary_text,
            ingredient_summary(primary_matches),
            ingredient_summary(secondary_matches),
            status,
            created_count,
            needs_review_count,
            score_basis,
            "; ".join(source_validation_flags),
            "; ".join(quality_flags),
            retailer_value(representative, "ingredients", max_len=5000),
            retailer_value(representative, "how_to_use", max_len=1500),
            retailer_value(representative, "product_attributes", max_len=1500),
        ] + [scores[col] for col in SCORE_COLUMNS]
        unique_rows.append(base_unique)

        if not retailer_rows:
            unmatched_products.append(
                [
                    auto_uid_base,
                    product["source_row"],
                    product["product_name"],
                    primary_text,
                    secondary_text,
                    status,
                ]
            )
            sku_retailer_rows = [None]
        else:
            sku_retailer_rows = retailer_rows

        for sku_index, retailer_row in enumerate(sku_retailer_rows, start=1):
            auto_uid = auto_uid_base if retailer_row is None else f"{auto_uid_base}-{retailer_value(retailer_row, 'site')}-{retailer_value(retailer_row, 'product_id') or sku_index}"
            sku_rows.append(
                [
                    auto_uid,
                    product["source_row"],
                    retailer_value(retailer_row, "id"),
                    retailer_value(retailer_row, "site"),
                    retailer_value(retailer_row, "parent_product_id"),
                    retailer_value(retailer_row, "product_id"),
                    retailer_value(retailer_row, "sku"),
                    retailer_value(retailer_row, "brand"),
                    product["product_name"],
                    retailer_value(retailer_row, "variant"),
                    category,
                    product_type,
                    safe_float(retailer_value(retailer_row, "mrp")) if retailer_row else "",
                    safe_float(retailer_value(retailer_row, "selling_price")) if retailer_row else "",
                    safe_float(retailer_value(retailer_row, "discount_pct")) if retailer_row else "",
                    safe_float(retailer_value(retailer_row, "rating")) if retailer_row else "",
                    safe_float(retailer_value(retailer_row, "rating_count")) if retailer_row else "",
                    safe_float(retailer_value(retailer_row, "review_count")) if retailer_row else "",
                    retailer_value(retailer_row, "in_stock"),
                    retailer_value(retailer_row, "product_url"),
                    retailer_value(retailer_row, "image_url"),
                    primary_text,
                    secondary_text,
                    ingredient_summary(primary_matches),
                    ingredient_summary(secondary_matches),
                    status,
                    created_count,
                    needs_review_count,
                    score_basis,
                    "; ".join(source_validation_flags),
                    "; ".join(quality_flags),
                    retailer_value(retailer_row, "key_ingredients"),
                    retailer_value(retailer_row, "ingredients", max_len=5000),
                    retailer_value(retailer_row, "how_to_use", max_len=1500),
                    retailer_value(retailer_row, "product_attributes", max_len=2000),
                ]
                + [scores[col] for col in SCORE_COLUMNS]
            )

    audit_headers = [
        "input_ingredient",
        "ingredient_group",
        "canonical_ingredient",
        "match_method",
        "confidence",
        "used_for_scoring",
        "fallback_family",
        "fallback_rationale",
        "product_examples",
    ]
    audit_rows = [
        [
            row["input_ingredient"],
            row["ingredient_group"],
            row["canonical_ingredient"],
            row["match_method"],
            row["confidence"],
            row["used_for_scoring"],
            row["fallback_family"],
            row["fallback_rationale"],
            " | ".join(row["product_examples"]),
        ]
        for row in sorted(match_audit_seen.values(), key=lambda item: (item["match_method"], item["input_ingredient"].lower()))
    ]

    ingredient_headers = [
        "Canonical Ingredient",
        "Source",
        "Skintype match / fallback family",
        "Concerns",
        "Fallback rationale",
    ] + SCORE_COLUMNS
    ingredient_rows = []
    extended_ingredients = original_ingredient_rows + list(fallback_rows.values())
    for row in extended_ingredients:
        ingredient_rows.append(
            [
                row["canonical_ingredient"],
                row["source"],
                row.get("skintype_match", ""),
                row.get("concerns", ""),
                row.get("fallback_rationale", ""),
            ]
            + [row["scores"].get(col, "") for col in SCORE_COLUMNS]
        )

    created_headers = ingredient_headers
    created_rows = [
        [
            row["canonical_ingredient"],
            row["source"],
            row.get("skintype_match", ""),
            row.get("concerns", ""),
            row.get("fallback_rationale", ""),
        ]
        + [row["scores"].get(col, "") for col in SCORE_COLUMNS]
        for row in fallback_rows.values()
    ]

    rules_headers = ["Area", "Option / Rule", "How it is handled"]
    rules_rows = [
        ["Skin type option", "Oily", "Uses Oily Score; if Sensitive=Yes uses Oily+Sensitive Score."],
        ["Skin type option", "Dry", "Uses Dry Score; if Sensitive=Yes uses Dry+Sensitive Score."],
        ["Skin type option", "Normal", "Uses Normal Score; if Sensitive=Yes uses Normal+Sensitive Score."],
        ["Skin type option", "Combination", "Uses Combination Score; if Sensitive=Yes uses Combination+Sensitive Score."],
        ["Skin type option", "I dont know", "New option from screenshot; uses average of base skin-type scores, or sensitive variants if Sensitive=Yes."],
        ["Sensitivity", "Yes / No", "Included through sensitive skin score columns."],
        ["Age", "Teen / Adult", "Teen maps to <16; Adult should average 17-25 and +>25 at profile-scoring time."],
        ["Gender", "Male / Female", "Gender itself is not scored; pregnancy and breastfeeding should remain disabled for male users."],
        ["Concern", "None", "New option from screenshot; uses None=100 so no-concern users are not penalized."],
        ["Concern label", "Uneven skin", "Mapped to Uneven Skin Tone."],
        ["Concern label", "Dark spots", "Mapped to Dark Spots/Pigmentation."],
        ["Concern label", "Wrinkles", "Mapped to Wrinkles/Fine lines."],
        ["Concern label", "redness", "Mapped to Redness/Irritation."],
        ["Concern group", "Face / Body", "Face concerns from screenshot are included; Body Acne retained from the ingredient score sheet."],
        ["Special condition", "Excessive dryness", "Uses Excessive Dryness score."],
        ["Special condition", "Pregnancy", "Uses Pregnancy Score; disabled in male quiz flow as shown."],
        ["Special condition", "Breast feeding", "Uses Breastfeeling Score; disabled in male quiz flow as shown."],
        ["Special condition", "None", "Uses None=100."],
        ["Product type", "Cleanser", "Skin type + concern are scored, except wrinkles where concern should be ignored at final profile scoring."],
        ["Product type", "Toner", "Skin type + concern are scored."],
        ["Product type", "Serum", "Primary ingredient group gets 80% weight and secondary group gets 20% per Scale sheet."],
        ["Product type", "Moisturizer / Sunscreen", "Skin type is scored; concern is not considered in final profile scoring per Scale sheet."],
        ["Product type", "Mask", "Skin type + concern are scored."],
        ["Hard blocker", "-100", "If any included ingredient has -100 for a score column, the product score for that column is -100."],
    ]

    total_useful_names = len(useful_products)
    matched_product_names = sum(1 for product in useful_products if product["product_name_key"] in retailer_by_name)
    matched_sku_rows = sum(len(retailer_by_name.get(product["product_name_key"], [])) for product in useful_products)
    unmatched_count = total_useful_names - matched_product_names
    summary_rows = [
        ["Metric", "Value"],
        ["Useful product names scored", total_useful_names],
        ["Useful product names matched to retailer metadata", matched_product_names],
        ["Useful product names missing retailer metadata", unmatched_count],
        ["Retailer SKU rows scored", matched_sku_rows],
        ["Scored SKU output rows including unmatched placeholders", len(sku_rows)],
        ["Original Roopsee ingredient score rows", len(original_ingredient_rows)],
        ["Created fallback ingredient rows", len(fallback_rows)],
        ["Extended ingredient master rows", len(ingredient_rows)],
        ["Unique ingredient audit rows", len(audit_rows)],
        ["Products with exact/alias ingredient matches only", status_counter["Matched to ingredient score sheet"]],
        ["Products with normalization but no created fallback", status_counter["Matched with normalization"]],
        ["Products with created fallback ingredients", status_counter["Mixed with created fallback"]],
        ["Products needing review because no scoreable ingredient remained", status_counter["Needs review - no scoreable ingredients"]],
        ["Top inferred product types", "; ".join(f"{name}: {count}" for name, count in type_counter.most_common(10))],
        ["Primary source workbook", str(USEFUL_PRODUCTS)],
        ["Retailer metadata CSV", str(RETAILER_PRODUCTS)],
        ["Ingredient scoring workbook", str(INGREDIENT_SCORES)],
    ]

    return {
        "summaryRows": summary_rows,
        "skuHeaders": sku_headers,
        "skuRows": sku_rows,
        "uniqueHeaders": unique_headers,
        "uniqueRows": unique_rows,
        "auditHeaders": audit_headers,
        "auditRows": audit_rows,
        "ingredientHeaders": ingredient_headers,
        "ingredientRows": ingredient_rows,
        "createdHeaders": created_headers,
        "createdRows": created_rows,
        "unmatchedHeaders": [
            "auto_product_uid",
            "source_row",
            "product_name",
            "primary_ingredients",
            "secondary_ingredients",
            "ingredient_match_status",
        ],
        "unmatchedRows": unmatched_products,
        "rulesHeaders": rules_headers,
        "rulesRows": rules_rows,
        "metrics": {
            "useful_product_names_scored": total_useful_names,
            "matched_product_names": matched_product_names,
            "unmatched_product_names": unmatched_count,
            "retailer_sku_rows_scored": matched_sku_rows,
            "scored_sku_rows_including_unmatched": len(sku_rows),
            "created_fallback_ingredient_rows": len(fallback_rows),
            "extended_ingredient_rows": len(ingredient_rows),
            "audit_rows": len(audit_rows),
            "product_type_counts": dict(type_counter),
            "status_counts": dict(status_counter),
        },
    }


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = build_rows()
    PAYLOAD_PATH.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(payload["metrics"], indent=2, ensure_ascii=False))
    print(PAYLOAD_PATH)


if __name__ == "__main__":
    main()
