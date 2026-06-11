from __future__ import annotations

import argparse
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from roopsee_coverage.constants import DEFAULT_PRODUCTS_CSV, DEFAULT_SCORE_WORKBOOK
from roopsee_coverage.engine import RecommendationEngine, coverage_status
from roopsee_coverage.profiles import (
    all_profile_combinations,
    concern_combinations,
    gender_free_special_states,
    skin_concern_special_combinations,
    skin_concern_type_combinations,
)


HEADERS = [
    "Profile ID",
    "Skin Type",
    "Age",
    "Concern Group",
    "Face & Body Concerns",
    "Lips & Eyes Concerns",
    "Special State",
    "Scoring Special Conditions",
    "Target Sheets",
    "Total Matching Products",
    "Products 90-100",
    "Products 80-89",
    "Products 70-79",
    "Products 60-69",
    "Products 50-59",
    "Products <50",
    "Products >=90",
    "Products >=80",
    "Products >=70",
    "Products >=60",
    "Products >=50",
    "Top Score",
    "Bottom Score",
    "Coverage Status",
    "Top 5 Products",
    "Profile JSON",
    "Scoring Profile JSON",
]


def band_counts(products: list[dict[str, Any]]) -> dict[str, int]:
    scores = [product["score"] for product in products]
    return {
        "90_100": sum(score >= 90 for score in scores),
        "80_89": sum(80 <= score < 90 for score in scores),
        "70_79": sum(70 <= score < 80 for score in scores),
        "60_69": sum(60 <= score < 70 for score in scores),
        "50_59": sum(50 <= score < 60 for score in scores),
        "below_50": sum(score < 50 for score in scores),
    }


def profile_row(profile_id: int, engine: RecommendationEngine, profile: dict[str, Any]) -> tuple[list[Any], str]:
    response = engine.recommend(profile, limit=250)
    products = response["products"]
    summary = response["summary"]
    bands = band_counts(products)
    status = coverage_status(response["profile"], summary)
    top_products = " | ".join(
        f"{product['product_name']} ({product['score']})" for product in products[:5]
    )
    row = [
        f"profile_{profile_id:05d}",
        profile["selectedSkinType"],
        profile["age"],
        profile.get("concernGroup", ""),
        ", ".join(profile.get("selectedFaceBodyConcerns", [])),
        ", ".join(profile.get("selectedLipsEyesConcerns", [])),
        profile.get("specialConditionState") or ", ".join(profile.get("selectedSpecialConditions", [])),
        ", ".join(response["profile"].get("selectedSpecialConditions", [])),
        ", ".join(response["target_sheets"]),
        response["total_matches"],
        bands["90_100"],
        bands["80_89"],
        bands["70_79"],
        bands["60_69"],
        bands["50_59"],
        bands["below_50"],
        summary["threshold_counts"]["above_90"],
        summary["threshold_counts"]["above_80"],
        summary["threshold_counts"]["above_70"],
        summary["threshold_counts"]["above_60"],
        summary["threshold_counts"]["above_50"],
        summary["top_score"],
        summary["bottom_score"],
        status,
        top_products,
        json.dumps(profile, ensure_ascii=False),
        json.dumps(response["profile"], ensure_ascii=False),
    ]
    return row, status


def style_sheet(ws, table_name: str) -> None:
    header_fill = PatternFill("solid", fgColor="263238")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = {
        "A": 14,
        "B": 16,
        "C": 14,
        "D": 16,
        "E": 34,
        "F": 34,
        "G": 30,
        "H": 28,
        "I": 18,
        "J": 16,
        "K": 15,
        "L": 15,
        "M": 15,
        "N": 15,
        "O": 15,
        "P": 13,
        "Q": 14,
        "R": 14,
        "S": 14,
        "T": 14,
        "U": 14,
        "V": 12,
        "W": 12,
        "X": 18,
        "Y": 64,
        "Z": 80,
        "AA": 80,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column >= 5)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_coverage_sheet(
    wb: Workbook,
    sheet_name: str,
    table_name: str,
    profiles: list[dict[str, Any]],
    engine: RecommendationEngine,
) -> dict[str, Any]:
    ws = wb.create_sheet(sheet_name)
    ws.append(HEADERS)
    status_counts: Counter[str] = Counter()
    for index, profile in enumerate(profiles, start=1):
        row, status = profile_row(index, engine, profile)
        ws.append(row)
        status_counts[status] += 1
    style_sheet(ws, table_name)
    return {
        "rows": len(profiles),
        "status_counts": dict(status_counts),
    }


def write_summary(wb: Workbook, sheet_results: dict[str, dict[str, Any]], assumptions: dict[str, Any]) -> None:
    ws = wb.create_sheet("Summary", 0)
    rows = [
        ("Section", "Metric", "Value"),
        ("PnC", "Formula", assumptions["formula"]),
        ("PnC", "All PnC rows", assumptions["all_pnc_rows"]),
        ("PnC", "Skin concern type rows", assumptions["skin_concern_type_rows"]),
        ("PnC", "With special conditions rows", assumptions["with_special_conditions_rows"]),
        ("Source", "Catalog products", assumptions["catalog_products"]),
        ("Source", "Catalog products missing score rows", assumptions["catalog_missing_score_count"]),
    ]
    for sheet_name, result in sheet_results.items():
        status_counts = result["status_counts"]
        rows.extend([
            (sheet_name, "Rows", result["rows"]),
            (sheet_name, "Strong profiles", status_counts.get("Strong", 0)),
            (sheet_name, "Usable profiles", status_counts.get("Usable", 0)),
            (sheet_name, "Limited but usable profiles", status_counts.get("Limited but usable", 0)),
            (sheet_name, "Coverage gap profiles", status_counts.get("Coverage gap", 0)),
        ])
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="263238")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 44
    ws.freeze_panes = "A2"


def write_assumptions(wb: Workbook, assumptions: dict[str, Any]) -> None:
    ws = wb.create_sheet("Assumptions")
    rows = [
        ("Rule", "Value"),
        ("Scoring basis", "Lowest applicable doctor-sheet score; no weighted or decimal score is generated."),
        ("Returned products", "Only products present in products.csv and matched by Product UID."),
        ("Gender", "Removed from final combination grid."),
        ("PnC formula", assumptions["formula"]),
        ("Skin profile", "4C1 = 4"),
        ("Concern formula", "2 * (8C1 + 8C2) = 2 * (8 + 28) = 72"),
        ("Special-condition formula", "3C0 + 3C1 + 3C2 + 3C3 + explicit None = 1 + 3 + 3 + 1 + 1 = 9"),
        ("Age factor", "4 states: Under 16, 17-25, Above 25, Not selected"),
        ("Age", "Under 16, 17-25, Above 25, and Not selected."),
        ("Not selected age handling", "Age score is skipped; it is not treated as Above 25."),
        ("Concern rule", "Face & Body or Lips & Eyes, choosing 1 or 2 concerns from that group."),
        ("Face & Body concern combinations", assumptions["face_concern_combinations"]),
        ("Lips & Eyes concern combinations", assumptions["lips_eye_concern_combinations"]),
        ("Special condition rule", "3C0 and explicit None are both listed as requested; both score as no special-condition constraint."),
        ("Special-condition combinations", assumptions["special_combinations"]),
        ("All PnC Combinations sheet", "Skin type + concern group/set + special state + age."),
        ("Skin Concern Type sheet", "Skin type + concern group/set; age and special conditions are not selected."),
        ("With Special Conditions sheet", "Skin type + concern group/set + special state; age is not selected."),
        ("All PnC rows", assumptions["all_pnc_rows"]),
        ("Skin concern type rows", assumptions["skin_concern_type_rows"]),
        ("With special conditions rows", assumptions["with_special_conditions_rows"]),
    ]
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="263238")
        cell.font = Font(color="FFFFFF", bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    ws.freeze_panes = "A2"
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_workbook(score_workbook: Path, products_csv: Path, output_path: Path) -> dict[str, Any]:
    engine = RecommendationEngine(score_workbook, products_csv)
    wb = Workbook()
    wb.remove(wb.active)

    all_profiles = all_profile_combinations(include_optional_age=True)
    skin_concern_profiles = skin_concern_type_combinations()
    special_profiles = skin_concern_special_combinations()

    sheet_results = {
        "All PnC Combinations": write_coverage_sheet(
            wb,
            "All PnC Combinations",
            "AllPnCCombinations",
            all_profiles,
            engine,
        ),
        "Skin Concern Type": write_coverage_sheet(
            wb,
            "Skin Concern Type",
            "SkinConcernType",
            skin_concern_profiles,
            engine,
        ),
        "With Special Conditions": write_coverage_sheet(
            wb,
            "With Special Conditions",
            "WithSpecialConditions",
            special_profiles,
            engine,
        ),
    }

    face_count = len([item for item in concern_combinations() if item["concern_group"] == "Face & Body"])
    lips_eye_count = len([item for item in concern_combinations() if item["concern_group"] == "Lips & Eyes"])
    health = engine.health()
    assumptions = {
        "formula": "4C1 * (2 * (8C1 + 8C2)) * (3C0 + 3C1 + 3C2 + 3C3 + None) * 4 = 10,368",
        "skin_types": 4,
        "age_states": 4,
        "face_concern_combinations": face_count,
        "lips_eye_concern_combinations": lips_eye_count,
        "concern_combinations": face_count + lips_eye_count,
        "special_combinations": len(gender_free_special_states()),
        "all_pnc_rows": len(all_profiles),
        "skin_concern_type_rows": len(skin_concern_profiles),
        "with_special_conditions_rows": len(special_profiles),
        "catalog_products": health["catalog_products"],
        "catalog_missing_score_count": health["catalog_missing_score_count"],
    }
    write_summary(wb, sheet_results, assumptions)
    write_assumptions(wb, assumptions)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "output_path": str(output_path),
        "sheet_results": sheet_results,
        "assumptions": assumptions,
        "health": health,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Export exhaustive Roopsee profile coverage workbook.")
    parser.add_argument("--score-workbook", default=DEFAULT_SCORE_WORKBOOK)
    parser.add_argument("--products-csv", default=DEFAULT_PRODUCTS_CSV)
    parser.add_argument("--output", default="outputs/Roopsee_Full_Profile_Coverage_Final.xlsx")
    args = parser.parse_args()

    result = export_workbook(
        Path(args.score_workbook).expanduser(),
        Path(args.products_csv).expanduser(),
        Path(args.output).expanduser(),
    )
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
