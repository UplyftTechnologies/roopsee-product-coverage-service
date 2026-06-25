from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK = BASE_DIR / "data" / "New products list 25062026.xlsx"
OUTPUT_CSV = BASE_DIR / "data" / "products.csv"

OUTPUT_COLUMNS = [
    "product_uid",
    "product_name",
    "brand_name",
    "category",
    "beat_the_heat",
    "gift_a_smile",
    "treat_yourself",
    "country_of_origin",
    "product_type",
    "addresses_skin_concerns",
    "sku_size",
    "mrp",
    "sp",
    "nykaa_brand_sp",
    "single_hero_ingredient",
    "secondary_hero_ingredients",
    "dos",
    "donts",
    "storage_instructions",
    "usage_instructions",
    "ingredient_cautions",
    "when_to_use",
    "ingredients",
    "product_description",
    "images",
    "created_at",
    "updated_at",
    "skin_type_not_available_comment",
    "id",
    "<16",
    "17-25",
    "Above 25",
    "Acne",
    "Body Acne",
    "Dryness",
    "Open Pores",
    "Uneven Skin Tone",
    "Dark Spots/Pigmentation",
    "Melasma",
    "Barrier Repair",
    "Comedones",
    "Wrinkles/Fine lines",
    "Redness/Irritation",
    "Dehydration",
    "Dullness",
    "Tanning",
    "Concern weight",
    "Oily Score",
    "Oily+Sensitive Score",
    "Dry Score",
    "Dry+Sensitive Score",
    "Normal Score",
    "Normal+Sensitive Score",
    "Combination Score",
    "Combination+Sensitive Score",
    "Excessive Dryness score",
    "Pregnancy Score",
    "Breastfeeling Score",
    "source_excel_row",
    "source_product_uid_status",
]

SCORE_COLUMNS = [
    "<16",
    "17-25",
    "Above 25",
    "Acne",
    "Body Acne",
    "Dryness",
    "Open Pores",
    "Uneven Skin Tone",
    "Dark Spots/Pigmentation",
    "Melasma",
    "Barrier Repair",
    "Comedones",
    "Wrinkles/Fine lines",
    "Redness/Irritation",
    "Dehydration",
    "Dullness",
    "Tanning",
    "Oily Score",
    "Oily+Sensitive Score",
    "Dry Score",
    "Dry+Sensitive Score",
    "Normal Score",
    "Normal+Sensitive Score",
    "Combination Score",
    "Combination+Sensitive Score",
    "Excessive Dryness score",
    "Pregnancy Score",
    "Breastfeeling Score",
]

FIELD_MAP = {
    "Product Name": "product_name",
    "Brand": "brand_name",
    "Category (L1)": "category",
    "Product Type (L2)": "product_type",
    "SKU": "sku_size",
    "MRP": "mrp",
    "Selling price ": "sp",
    "Country of Origin": "country_of_origin",
    "Concerns": "addresses_skin_concerns",
    "Single Hero Ingredient": "single_hero_ingredient",
    "Secondary Hero Ingredients": "secondary_hero_ingredients",
    "Product description": "product_description",
    "Ingredients": "ingredients",
    "Do's": "dos",
    "Donts": "donts",
    "Storage Instructions": "storage_instructions",
    "Usage Instructions": "usage_instructions",
    "Images": "images",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u00a0", " ").split()).strip()


def clean_number(value: Any) -> str:
    if value is None or clean_text(value) == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug[:72] or "product"


def canonical_product_type(value: Any) -> str:
    raw = clean_text(value)
    normalized = raw.lower()
    if normalized == "seru":
        return "Serum"
    if normalized == "moisturiser":
        return "Moisturizer"
    return raw


def has_any(text: str, needles: list[str]) -> bool:
    lowered = text.lower()
    return any(needle in lowered for needle in needles)


def infer_when_to_use(row: dict[str, str]) -> str:
    product_type = row["product_type"].lower()
    text = " ".join(
        [
            row["product_name"],
            row["product_type"],
            row["single_hero_ingredient"],
            row["secondary_hero_ingredients"],
            row["ingredients"],
            row["usage_instructions"],
        ]
    )

    if "sunscreen" in product_type or has_any(text, ["spf", "sunscreen", "uv filter"]):
        return "Morning"
    if has_any(text, ["retinol", "retinal", "retinoid", "tretinoin", "adapalene"]):
        return "Night"
    if has_any(text, ["sleeping mask", "overnight"]):
        return "Night"
    if product_type in {"serum", "toner", "mask", "treatment"} and has_any(
        text,
        [
            "glycolic",
            "lactic acid",
            "mandelic",
            "salicylic",
            "bha",
            "aha",
            "pha",
            "peeling",
            "exfoliating",
        ],
    ):
        return "Night"
    if product_type in {"cleanser", "wash", "body wash", "moisturizer", "body lotion", "lotion", "cream"}:
        return "Morning and Night"
    if product_type == "serum":
        return "Morning and Night"
    return "Morning and Night"


def import_products() -> tuple[int, int, int]:
    workbook = load_workbook(SOURCE_WORKBOOK, data_only=True)
    worksheet = workbook["Sheet2"]
    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]

    raw_rows: list[dict[str, Any]] = []
    for row_number in range(2, worksheet.max_row + 1):
        raw = {
            headers[column - 1]: worksheet.cell(row_number, column).value
            for column in range(1, worksheet.max_column + 1)
            if headers[column - 1] is not None
        }
        if not any(value not in (None, "") for value in raw.values()):
            continue
        raw["_source_excel_row"] = row_number
        raw_rows.append(raw)

    provided_counts: dict[str, int] = {}
    for raw in raw_rows:
        uid = clean_text(raw.get("Product UID"))
        if uid:
            provided_counts[uid] = provided_counts.get(uid, 0) + 1

    seen_uids: dict[str, int] = {}
    rows: list[dict[str, str]] = []
    generated_count = 0
    duplicate_count = 0
    for raw in raw_rows:
        output = {column: "" for column in OUTPUT_COLUMNS}
        source_row = int(raw["_source_excel_row"])
        raw_uid = clean_text(raw.get("Product UID"))
        name = clean_text(raw.get("Product Name"))
        brand = clean_text(raw.get("Brand"))

        if not raw_uid:
            generated_count += 1
            uid = f"generated-25062026-row-{source_row}-{slugify(name)}"
            status = "generated_missing_uid"
        elif raw_uid in seen_uids:
            duplicate_count += 1
            seen_uids[raw_uid] += 1
            uid = f"{raw_uid}-duplicate-row-{source_row}"
            status = "generated_duplicate_uid"
        else:
            seen_uids[raw_uid] = 1
            uid = raw_uid
            status = "provided"

        output["product_uid"] = uid
        output["source_product_uid_status"] = status
        output["source_excel_row"] = str(source_row)
        output["id"] = str(source_row)
        output["brand_name"] = brand
        output["nykaa_brand_sp"] = clean_number(raw.get("Selling price ")) or clean_number(raw.get("MRP"))
        output["sp"] = clean_number(raw.get("Selling price ")) or clean_number(raw.get("MRP"))

        for source, target in FIELD_MAP.items():
            if target == "product_type":
                output[target] = canonical_product_type(raw.get(source))
            elif target in {"mrp"}:
                output[target] = clean_number(raw.get(source))
            else:
                output[target] = clean_text(raw.get(source))

        for score_column in SCORE_COLUMNS:
            output[score_column] = clean_number(raw.get(score_column))

        output["when_to_use"] = infer_when_to_use(output)
        rows.append(output)

    with OUTPUT_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)

    return len(rows), generated_count, duplicate_count


if __name__ == "__main__":
    total, generated, duplicates = import_products()
    print(f"Imported {total} products into {OUTPUT_CSV}")
    print(f"Generated missing UIDs: {generated}")
    print(f"Generated duplicate UIDs: {duplicates}")
